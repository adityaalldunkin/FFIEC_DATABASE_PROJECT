#!/usr/bin/env python3
"""
Build texas_loans_summary.xlsx — 5-sheet borrower workbook.

  1. texas_loans_summary  — original regulatory loan line items (CSV export)
  2. EDA                  — exhaustive exploratory analysis (borrower lens)
  3. Insights             — documented findings with plain-English explanations
  4. Data Provenance      — how the data was retrieved (FFIEC API pipeline)
  5. Abbreviations        — glossary of every abbreviation in the dataset & analysis

  python ONLY_TEXAS_SINCE_2025/build_texas_loans_summary_borrower_eda.py
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent
REPO = ROOT.parent
INPUT = ROOT / "exports" / "texas_loans_summary.csv"
OUTPUT = REPO / "texas_loans_summary.xlsx"
OUTPUT_COPY = ROOT / "analysis" / "texas_loans_summary.xlsx"
PROFILES = ROOT / "exports" / "texas_bank_profiles_latest.csv"
FILINGS = ROOT / "exports" / "texas_filings.csv"
INSTITUTIONS = ROOT / "exports" / "texas_institutions.csv"
PROGRESS = ROOT / "data" / "progress.json"
TAXONOMY = REPO / "texas_mdrm_loan_taxonomy.csv"

ICP_MIN = 500_000_000
ICP_MAX = 2_000_000_000
SPECIALIST_PCT = 8.0

PRODUCT_LINES = {
    "total_loans_gross": ["RCON2122", "RCFD2122"],
    "multifamily_re_loans": ["RCON1460"],
    "other_nonfarm_nonres_re": ["RCONF161"],
    "commercial_re_loans": ["RCONF162"],
    "owner_occupied_nonfarm_re": ["RCONF160"],
    "residential_construction": ["RCONF158"],
    "other_construction_ld": ["RCONF159"],
    "residential_1_4_family": ["RCON1403"],
    "credit_card_plans": ["RCON1545"],
    "other_consumer_loans": ["RCON1583"],
    "farmland_loans": ["RCON1420"],
    "ag_production_loans": ["RCON1590"],
    "lease_financing": ["RCON1754", "RCONF163"],
    "revolving_1_4_family": ["RCON1797"],
    "residential_first_lien": ["RCON5367"],
    "residential_junior_lien": ["RCON5368"],
    "ci_loans": [],
}

MIX_DEF = {
    "mf": ("Multifamily (5+ units)", ["multifamily_re_loans"]),
    "inv": ("Investor CRE (income property)", ["other_nonfarm_nonres_re", "commercial_re_loans"]),
    "own": ("Owner-Occupied CRE", ["owner_occupied_nonfarm_re"]),
    "con": ("Commercial Construction", ["residential_construction", "other_construction_ld"]),
    "ci": ("Commercial & Industrial (C&I)", ["ci_loans"]),
    "res": ("1–4 Family Residential", ["residential_1_4_family"]),
    "heloc": ("HELOC / revolving 1–4 family", ["revolving_1_4_family"]),
    "cons": ("Consumer lending", ["credit_card_plans", "other_consumer_loans"]),
    "farm": ("Agricultural & Farmland (farmland)", ["farmland_loans"]),
    "ag": ("Agricultural & Farmland (production)", ["ag_production_loans"]),
    "lease": ("Lease financing", ["lease_financing"]),
}

CODE_TO_BORROWER = [
    ("RCON2122 / RCFD2122", "Total loans & leases (headline portfolio size)", "All borrowers — denominator for mix %"),
    ("RCONF161", "Investor CRE (income property)", "Acquisition, bridge, permanent CRE"),
    ("RCONF160", "Owner-occupied CRE", "User-owned commercial real estate"),
    ("RCONF158 / RCONF159", "Construction & land development", "Ground-up, value-add, land loans"),
    ("RCON1460", "Multifamily (5+ units)", "Apartment acquisition, refi, bridge"),
    ("RCON1403", "1–4 family residential", "Home mortgages (aggregated)"),
    ("RCON1797", "HELOC / revolving 1–4 family", "Home equity lines"),
    ("RCON1420 / RCON1590", "Farmland & ag production", "Farm operating and land loans"),
    ("RCON1545 / RCON1583", "Credit cards & other consumer", "Retail / consumer (not CRE/C&I)"),
    ("RCON1754 / RCONF163", "Lease financing", "Equipment / commercial lease"),
    ("RCON1766 (profiles)", "C&I / business loans", "Operating lines, term loans, working capital"),
    ("RCON5367 / RCON5368", "Residential 1–4 lien balances", "Mortgage market depth signal"),
]

EXPLANATION_FILL = PatternFill("solid", fgColor="F2F2F2")
EXPLANATION_LABEL_FONT = Font(bold=True, size=10, color="1F4E79")


def round_eda_table(df: pd.DataFrame) -> pd.DataFrame:
    """Round every numeric cell to the nearest whole number for EDA display."""
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].apply(lambda x: int(round(x)) if pd.notna(x) else x)
    return out


def _fmt_usd_m(x: float) -> int:
    return int(round(x / 1e6))


def _fmt_usd_b(x: float) -> int:
    return int(round(x / 1e9))


def _explain_specialists(label: str, top: pd.DataFrame, n_specialists: int) -> str:
    if top.empty:
        return f"No Texas banks reported a non-zero {label} balance in the latest quarter."
    lead = top.iloc[0]
    bank = lead.get("Bank", lead.get("institution_name", "N/A"))
    pct = lead.get("Portfolio %", 0)
    return (
        f"This table ranks the ten Texas banks with the highest share of their loan book in {label}. "
        f"{bank} leads at {int(round(pct))}% portfolio concentration. "
        f"Statewide, {n_specialists} banks meet the ≥{int(SPECIALIST_PCT)}% specialist threshold for this product. "
        f"Banks at the top of this list have made {label} a core line of business — they are the strongest first-call "
        f"targets for borrowers in that category. Portfolio % shows product balance divided by total loans (RCON2122); "
        f"higher values mean the bank is more likely to understand and price your deal type."
    )

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=12)
BANNER_FILL = PatternFill("solid", fgColor="2E75B6")
BANNER_FONT = Font(bold=True, color="FFFFFF", size=11)
INSIGHT_HIGH = PatternFill("solid", fgColor="E2EFDA")
INSIGHT_MED = PatternFill("solid", fgColor="FFF2CC")
PROV_SECTION = PatternFill("solid", fgColor="E7E6E6")
PROV_SECTION_FONT = Font(bold=True, size=11)


def parse_periods(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["period_dt"] = pd.to_datetime(out["reporting_period"], dayfirst=False)
    return out


def gini(values: np.ndarray) -> float:
    x = np.sort(values[values > 0])
    if len(x) == 0:
        return float("nan")
    n = len(x)
    cum = np.cumsum(x)
    return float((n + 1 - 2 * np.sum(cum) / cum[-1]) / n)


def hhi(shares: np.ndarray) -> float:
    s = shares[shares > 0]
    if len(s) == 0:
        return float("nan")
    total = s.sum()
    if total <= 0:
        return float("nan")
    p = s / total
    return float((p ** 2).sum())


def form_segment(form: str) -> str:
    f = str(form or "")
    if "041" in f:
        return "Community bank (FFIEC 041)"
    if "031" in f or "032" in f:
        return "Regional / larger (FFIEC 031/032)"
    return "Other / specialty form"


def pick_line_value(group: pd.DataFrame, col: str) -> float:
    codes = PRODUCT_LINES[col]
    if not codes:
        return float(group[col].iloc[0]) if col in group.columns else 0.0
    sub = group[group["mdrm_code"].isin(codes)]
    if sub.empty:
        return 0.0
    return float(sub.sort_values("mdrm_code")["value_num"].max())


def load_profiles() -> pd.DataFrame:
    if not PROFILES.exists():
        return pd.DataFrame()
    p = pd.read_csv(PROFILES, dtype={"id_rssd": int})
    p["assets_usd"] = p["total_assets"]
    p["icp_fit"] = p.get("icp_fit", pd.Series(dtype=str))
    p["is_icp"] = p["icp_fit"] == "ICP ($500M–$2B)"
    if p["is_icp"].sum() == 0:
        p["is_icp"] = p["assets_usd"].between(ICP_MIN, ICP_MAX)
    return p


def build_bank_wide(period_df: pd.DataFrame, profiles: pd.DataFrame) -> pd.DataFrame:
    prof_ci = {}
    if not profiles.empty and "ci_loans" in profiles.columns:
        prof_ci = profiles.set_index("id_rssd")["ci_loans"].to_dict()
    meta_cols = ["id_rssd", "institution_name", "reporting_form", "reporting_period"]
    rows = []
    for rssd, grp in period_df.groupby("id_rssd"):
        row = {c: grp[c].iloc[0] for c in meta_cols}
        for col in PRODUCT_LINES:
            row[col] = pick_line_value(grp, col)
        row["ci_loans"] = float(prof_ci.get(rssd, 0.0))
        rows.append(row)
    return pd.DataFrame(rows)


def compute_mix_frame(wide: pd.DataFrame) -> pd.DataFrame:
    out = wide.copy()
    total = out["total_loans_gross"].clip(lower=0)
    out["total_loans_usd"] = total
    for key, (_, parts) in MIX_DEF.items():
        bal = sum(out[p].fillna(0) for p in parts)
        out[f"mix_{key}_usd"] = bal
        out[f"mix_{key}_pct"] = np.where(total > 0, 100 * bal / total, 0)
    accounted = sum(out[f"mix_{k}_usd"] for k in MIX_DEF)
    out["mix_uncat_usd"] = np.maximum(0, total - accounted)
    out["mix_uncat_pct"] = np.where(total > 0, 100 * out["mix_uncat_usd"] / total, 0)
    out["form_segment"] = out["reporting_form"].map(form_segment)
    return out


def enrich_mix(mix: pd.DataFrame, profiles: pd.DataFrame) -> pd.DataFrame:
    if not profiles.empty:
        cols = ["id_rssd", "city", "total_assets", "is_icp", "icp_fit", "past_due_90_plus"]
        cols = [c for c in cols if c in profiles.columns]
        mix = mix.merge(profiles[cols].drop_duplicates("id_rssd"), on="id_rssd", how="left")
    else:
        mix["city"] = ""
        mix["total_assets"] = np.nan
        mix["is_icp"] = mix["total_loans_usd"].between(ICP_MIN, ICP_MAX)
        mix["past_due_90_plus"] = 0
    mix["past_due_90_plus_pct"] = np.where(
        mix["total_loans_usd"] > 0,
        100 * mix.get("past_due_90_plus", pd.Series(0, index=mix.index)).fillna(0) / mix["total_loans_usd"],
        0,
    )
    mix["portfolio_style"] = (
        (mix["mix_inv_pct"] + mix["mix_own_pct"] + mix["mix_ci_pct"]) > 50
    ).map({True: "CRE+C&I portfolio", False: "Other mix"})
    return mix


def run_analysis(df: pd.DataFrame) -> tuple[list, list]:
    df = parse_periods(df)
    profiles = load_profiles()
    latest_period = df["period_dt"].max()
    latest = df[df["period_dt"] == latest_period].copy()
    wide = build_bank_wide(latest, profiles)
    mix = enrich_mix(compute_mix_frame(wide), profiles)
    active = mix[mix["total_loans_gross"] > 0].copy()
    n_banks = len(active)
    period_str = latest_period.strftime("%Y-%m-%d")

    eda: list[tuple[str, object]] = []
    insights: list[dict] = []

    def section(title: str) -> None:
        eda.append((title, ""))
        eda.append(("", ""))

    def table(df_out: pd.DataFrame, explanation: str) -> None:
        eda.append(("__table__", round_eda_table(df_out)))
        eda.append(("__explanation__", explanation))
        eda.append(("", ""))

    def insight(**kwargs) -> None:
        insights.append(kwargs)

    # ── Banner ──────────────────────────────────────────────────────────────
    section("Borrower-perspective EDA — texas_loans_summary")
    eda.extend([
        ("Analysis date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Latest reporting period", period_str),
        ("Reporting periods in file", int(df["period_dt"].nunique())),
        ("Distinct Texas banks (latest)", int(n_banks)),
        ("Total rows in source file", f"{len(df):,}"),
        ("Distinct MDRM codes", int(df["mdrm_code"].nunique())),
        ("Value units", "USD (aligned with texas_bank_profiles_latest.csv)"),
        ("C&I enrichment", "RCON1766 merged from bank profiles (not in summary extract)"),
        ("Analytical lens", "Help borrowers find the right Texas bank by product specialization, market depth, geography, and lender health"),
        ("", ""),
    ])

    # ── 0. Column guide ─────────────────────────────────────────────────────
    section("0. Source file column guide")
    table(pd.DataFrame([
        {"Column": "id_rssd", "Meaning": "Federal Reserve RSSD ID — stable bank identifier"},
        {"Column": "institution_name", "Meaning": "Legal/reporting name from FFIEC panel"},
        {"Column": "reporting_period", "Meaning": "Call Report quarter end (MM/DD/YY)"},
        {"Column": "mdrm_code", "Meaning": "Regulatory line code (e.g. RCON2122, RCONF161)"},
        {"Column": "item_name", "Meaning": "Federal Reserve MDRM short label"},
        {"Column": "mdrm_description", "Meaning": "Full Fed definition of the line item"},
        {"Column": "value_num", "Meaning": "Reported balance in USD for that category"},
        {"Column": "reporting_form", "Meaning": "FFIEC 031 (larger) or 041 (community) form"},
    ]), (
        "Every row in the source file is one regulatory loan line item for one bank in one quarter. "
        "Borrowers should focus on id_rssd (to identify the bank), mdrm_code (the product category), "
        "and value_num (the dollar balance). The MDRM codes are how regulators classify loans — "
        "they do not match marketing names on bank websites, but they are the authoritative measure of "
        "what each bank actually holds on its balance sheet."
    ))

    # ── 1. MDRM → borrower mapping ──────────────────────────────────────────
    section("1. Regulatory code → borrower product mapping")
    table(pd.DataFrame(CODE_TO_BORROWER, columns=["MDRM code(s)", "Borrower product", "Typical use case"]), (
        "This crosswalk translates opaque regulatory codes into borrower-friendly product names. "
        "When outreach to a bank, match your deal type to the MDRM line they report most heavily — "
        "for example, investor acquisitions map to RCONF161, operating lines map to RCON1766 (C&I), "
        "and apartment deals map to RCON1460. Use this table to know which code to filter on in the raw data."
    ))

    # ── 2. Availability ───────────────────────────────────────────────────────
    section("2. Lender availability by borrower product (latest quarter)")
    avail_rows = []
    for key, (label, _) in MIX_DEF.items():
        bal_col = f"mix_{key}_usd"
        pct_col = f"mix_{key}_pct"
        lenders_any = int((active[bal_col] > 0).sum())
        specialists = int((active[pct_col] >= SPECIALIST_PCT).sum())
        deep = int((active[pct_col] >= 15).sum())
        total_mkt = active[bal_col].sum()
        med = active.loc[active[bal_col] > 0, bal_col].median() if lenders_any else 0
        avail_rows.append({
            "Borrower product": label,
            "Banks with any balance": lenders_any,
            "Banks ≥ $1M": int((active[bal_col] >= 1e6).sum()),
            "Banks ≥ $10M": int((active[bal_col] >= 10e6).sum()),
            f"Specialists ≥{int(SPECIALIST_PCT)}%": specialists,
            "Deep specialists ≥15%": deep,
            "TX market ($B)": int(round(total_mkt / 1e9)),
            "Median active lender ($M)": int(round(med / 1e6)) if lenders_any else 0,
            "Availability %": int(round(100 * lenders_any / n_banks)),
        })
    avail_df = pd.DataFrame(avail_rows)
    table(avail_df, (
        f"C&I and CRE categories show the broadest lender availability — most Texas banks hold some balance in these lines. "
        f"Consumer lending and lease financing are far narrower: only a small fraction of banks are meaningful players. "
        f"The 'Specialists' column counts banks where the product is ≥{int(SPECIALIST_PCT)}% of total loans — your best outreach targets. "
        f"'TX market ($B)' is the aggregate balance statewide; 'Median active lender ($M)' shows the typical bank's exposure "
        f"among those that participate. High availability % means you can build a competitive shortlist; low availability "
        f"means you must be selective and relationship-driven."
    ))

    best = avail_df.loc[avail_df["Availability %"].idxmax()]
    worst = avail_df.loc[avail_df["Availability %"].idxmin()]
    insight(
        id="I-01", priority="High", theme="Product choice breadth", section="EDA §2",
        insight=(
            f"{best['Borrower product']} has the widest lender pool ({int(best['Banks with any balance'])} of {n_banks} banks, "
            f"{best['Availability %']:.0f}% availability). {worst['Borrower product']} is the narrowest "
            f"({int(worst['Banks with any balance'])} banks, {worst['Availability %']:.0f}%)."
        ),
        explanation=(
            "Availability score measures how many Texas banks report a non-zero balance in your product category. "
            "High availability means you can shop multiple lenders; low availability means fewer realistic options "
            "and relationship quality matters more than volume outreach."
        ),
        evidence=f"Latest period {period_str}; specialist threshold {SPECIALIST_PCT}% of total loans.",
        borrower_action="Prioritize high-availability categories for competitive terms; for niche products, target specialists only.",
    )

    # ── 3. Concentration ──────────────────────────────────────────────────────
    section("3. Market concentration (HHI & Gini) — how competitive is each product?")
    conc_rows = []
    for key, (label, _) in MIX_DEF.items():
        bal = active[f"mix_{key}_usd"].values
        h = hhi(bal)
        g = gini(bal)
        top5 = float(np.sort(bal)[-5:].sum() / bal.sum() * 100) if bal.sum() > 0 else 0
        conc_rows.append({
            "Borrower product": label,
            "HHI × 10,000": int(round(h * 10000)),
            "Gini × 100": int(round(g * 100)),
            "Top 5 share %": int(round(top5)),
            "Label": "Highly concentrated" if h > 0.15 else "Moderate" if h > 0.08 else "Fragmented",
        })
    conc_df = pd.DataFrame(conc_rows)
    table(conc_df, (
        "HHI × 10,000 is the standard antitrust scale (higher = more concentrated). Consumer lending shows the highest "
        "concentration — a handful of large banks dominate. C&I and CRE are more fragmented, giving borrowers more choice. "
        "Gini measures inequality of balances across banks (0 = equal, 100 = one bank holds everything). "
        "'Top 5 share %' shows what fraction of the statewide product market the five largest banks control. "
        "In fragmented markets you can negotiate across many lenders; in concentrated markets you still need alternatives "
        "but should include mid-size community banks that may compete on terms."
    ))

    top3 = conc_df.nlargest(3, "HHI × 10,000")
    insight(
        id="I-02", priority="High", theme="Market structure", section="EDA §3",
        insight="Most concentrated markets: " + "; ".join(
            f"{r['Borrower product']} (top 5 = {r['Top 5 share %']:.0f}%)" for _, r in top3.iterrows()
        ) + ".",
        explanation=(
            "HHI (Herfindahl-Hirschman Index) measures how much of a product market is held by a few banks. "
            "Above 0.15 is 'highly concentrated' — borrowers may have limited alternatives. Below 0.08 is fragmented — "
            "more banks compete, which can improve terms if you shop."
        ),
        evidence="HHI computed on bank-level product balances, latest quarter.",
        borrower_action="In concentrated markets, still build an 8–10 bank list including mid-size community banks.",
    )

    # ── 4. Exposure distribution ────────────────────────────────────────────
    section("4. Exposure distribution among active lenders (P25 / median / P75)")
    dist_rows = []
    for key, (label, _) in MIX_DEF.items():
        col = f"mix_{key}_usd"
        pos = active.loc[active[col] > 0, col]
        if pos.empty:
            continue
        dist_rows.append({
            "Product": label,
            "P25 ($M)": int(round(pos.quantile(0.25) / 1e6)),
            "Median ($M)": int(round(pos.median() / 1e6)),
            "P75 ($M)": int(round(pos.quantile(0.75) / 1e6)),
            "Max ($M)": int(round(pos.max() / 1e6)),
        })
    table(pd.DataFrame(dist_rows), (
        "These percentiles show how large each product line is at the typical active Texas bank. "
        "Compare your requested loan size to the median and P75: if your deal is above the median, it is material "
        "to that bank and they are more likely to engage. If your deal is far below P25, you may be too small for "
        "their committee process unless you offer relationship value. The max column shows the largest single-bank "
        "exposure in each category — useful for understanding who can hold big tickets."
    ))

    insight(
        id="I-03", priority="Medium", theme="Deal size fit", section="EDA §4",
        insight=(
            "Median active-lender exposures span orders of magnitude. A $3M CRE loan is material for many community "
            "banks but small for regional portfolios."
        ),
        explanation=(
            "Compare your requested loan size to the median product balance of banks you target. Banks where your deal "
            "is above their median exposure in that product are more likely to engage — it is meaningful to their book."
        ),
        evidence="Percentiles of positive product balances per category.",
        borrower_action="Disclose loan size early; ask if it clears internal hold limits.",
    )

    # ── 5. Top specialists (all products) ───────────────────────────────────
    section("5. Top 10 specialist lenders by product (portfolio share %)")
    for key, (label, _) in MIX_DEF.items():
        pct_col, usd_col = f"mix_{key}_pct", f"mix_{key}_usd"
        n_spec = int((active[pct_col] >= SPECIALIST_PCT).sum())
        top = active[active[usd_col] > 0].nlargest(10, pct_col)[
            ["institution_name", "city", pct_col, usd_col, "total_loans_usd", "form_segment"]
        ].copy()
        if top.empty:
            continue
        top.columns = ["Bank", "City", "Portfolio %", "Product ($M)", "Total loans ($M)", "Segment"]
        top["Portfolio %"] = top["Portfolio %"].round(0).astype(int)
        top["Product ($M)"] = (top["Product ($M)"] / 1e6).round(0).astype(int)
        top["Total loans ($M)"] = (top["Total loans ($M)"] / 1e6).round(0).astype(int)
        eda.append((f"Top specialists — {label}", ""))
        table(top.reset_index(drop=True), _explain_specialists(label, top, n_spec))

    inv_top = active.nlargest(3, "mix_inv_pct")
    insight(
        id="I-04", priority="Medium", theme="Specialist targeting", section="EDA §5",
        insight="Top investor CRE specialists: " + ", ".join(
            f"{r.institution_name} ({r.mix_inv_pct:.0f}%, {r.city or 'TX'})" for _, r in inv_top.iterrows()
        ) + ".",
        explanation=(
            "A bank with ≥8% of its loan book in your product category is a 'specialist' — they have demonstrated appetite. "
            "Above 15% is a deep specialist. These banks are your highest-probability first calls for that product."
        ),
        evidence="RCONF161 + RCONF162 → Investor CRE; ranked by mix_inv_pct.",
        borrower_action="Lead with NOI, occupancy, and sponsor track record for CRE outreach.",
    )

    # ── 6. Top banks by total loans ─────────────────────────────────────────
    section("6. Top 20 Texas banks by total loan portfolio (latest quarter)")
    top20 = active.nlargest(20, "total_loans_usd")[
        ["institution_name", "city", "total_loans_usd", "mix_inv_pct", "mix_ci_pct", "mix_mf_pct", "form_segment"]
    ].copy()
    top20.columns = ["Bank", "City", "Total loans ($M)", "Investor CRE %", "C&I %", "MF %", "Segment"]
    top20["Total loans ($M)"] = top20["Total loans ($M)"].apply(
        lambda x: int(round(x / 1e6)) if x < 1e9 else int(round(x / 1e6))
    )
    for c in ["Investor CRE %", "C&I %", "MF %"]:
        top20[c] = top20[c].round(0).astype(int)
    table(top20.reset_index(drop=True), (
        "These are the twenty largest loan portfolios among Texas banks in the latest quarter. "
        "Size alone does not mean fit — check the mix columns to see whether each bank is CRE-heavy, C&I-heavy, "
        "or multifamily-focused. Regional and money-center banks dominate by total loans but may be slower and "
        "more price-driven; community banks further down the list may offer better relationship terms for "
        "mid-market commercial borrowers."
    ))

    # ── 7. Form segmentation ────────────────────────────────────────────────
    section("7. Community bank (FFIEC 041) vs regional (031/032)")
    seg_rows = []
    for seg in sorted(active["form_segment"].unique()):
        sub = active[active["form_segment"] == seg]
        seg_rows.append({
            "Segment": seg,
            "Banks": len(sub),
            "Median loans ($M)": int(round(sub["total_loans_usd"].median() / 1e6)),
            "Median C&I %": int(round(sub["mix_ci_pct"].median())),
            "Median inv CRE %": int(round(sub["mix_inv_pct"].median())),
            "Portfolio-style (CRE+C&I>50%)": int(((sub["mix_inv_pct"] + sub["mix_own_pct"] + sub["mix_ci_pct"]) > 50).sum()),
        })
    table(pd.DataFrame(seg_rows), (
        "FFIEC 041 filers are community banks — smaller median portfolios and often higher CRE/C&I mix relative to size. "
        "FFIEC 031/032 filers are larger regionals with bigger balance sheets. The 'Portfolio-style' count shows how many "
        "banks in each segment are primarily commercial lenders (CRE + C&I > 50% of loans). Commercial borrowers should "
        "weight outreach toward 041 portfolio-style banks for relationship-driven deals and faster credit decisions."
    ))

    comm = active[active["form_segment"] == "Community bank (FFIEC 041)"]
    insight(
        id="I-05", priority="High", theme="Bank size fit", section="EDA §7",
        insight=(
            f"{len(comm)} community banks (FFIEC 041); median portfolio ${comm['total_loans_usd'].median()/1e6:.0f}M. "
            f"{int(((comm['mix_inv_pct']+comm['mix_own_pct']+comm['mix_ci_pct'])>50).sum())} are portfolio-style lenders."
        ),
        explanation=(
            "FFIEC 041 filers are typically community banks with simpler structures and faster credit committees. "
            "Portfolio-style banks (CRE + C&I > 50% of loans) are the best fit for commercial borrowers — "
            "they are not retail mortgage or credit-card shops."
        ),
        evidence="reporting_form field; portfolio-style = inv + own + ci > 50%.",
        borrower_action="Prefer 041 banks for relationship-driven CRE/C&I deals.",
    )

    # ── 8. ICP band ─────────────────────────────────────────────────────────
    section("8. $500M–$2B asset band (ICP community banks)")
    icp = active[active["is_icp"] == True]  # noqa: E712
    non_icp = active[active["is_icp"] != True]  # noqa: E712
    table(pd.DataFrame([
        {"Segment": "ICP ($500M–$2B)", "Banks": len(icp),
         "Median loans ($M)": int(round(icp["total_loans_usd"].median() / 1e6)) if len(icp) else 0,
         "Median CRE %": int(round((icp["mix_inv_pct"] + icp["mix_own_pct"]).median())) if len(icp) else 0,
         "Median C&I %": int(round(icp["mix_ci_pct"].median())) if len(icp) else 0},
        {"Segment": "Outside ICP", "Banks": len(non_icp),
         "Median loans ($M)": int(round(non_icp["total_loans_usd"].median() / 1e6)),
         "Median CRE %": int(round((non_icp["mix_inv_pct"] + non_icp["mix_own_pct"]).median())),
         "Median C&I %": int(round(non_icp["mix_ci_pct"].median()))},
    ]), (
        f"The ICP ($500M–$2B asset) band contains {len(icp)} banks — the sweet spot for commercial borrowers seeking "
        f"$2M–$25M relationship credits. These banks are large enough to hold meaningful exposure but small enough to "
        f"value direct sponsor relationships. Compare median CRE % and C&I % to see that ICP banks are not dominated "
        f"by retail consumer lending, making them better targets than very small banks (limited capacity) or very "
        f"large banks (transactional pricing)."
    ))
    icp_spec = {MIX_DEF[k][0]: int((icp[f"mix_{k}_pct"] >= SPECIALIST_PCT).sum()) if len(icp) else 0 for k in MIX_DEF}
    eda.append(("ICP specialists by product (≥8%)", ""))
    table(
        pd.DataFrame([{"Product": k, "ICP specialists": v} for k, v in sorted(icp_spec.items(), key=lambda x: -x[1])]),
        (
            "This table counts how many ICP-band banks are specialists (≥8% portfolio share) in each product. "
            "High counts in C&I and Investor CRE mean ICP borrowers have many credible options without going to "
            "money-center banks. Low counts in Multifamily and Construction mean niche products require more "
            "targeted research even within the ICP segment."
        ),
    )

    section("8b. ICP bank shortlist — top 30 by investor CRE portfolio share")
    icp_list = icp[icp["mix_inv_usd"] > 0].nlargest(30, "mix_inv_pct")[
        ["institution_name", "city", "total_assets", "mix_inv_pct", "mix_inv_usd", "mix_ci_pct"]
    ].copy()
    icp_list.columns = ["Bank", "City", "Assets ($M)", "Investor CRE %", "Inv CRE ($M)", "C&I %"]
    icp_list["Assets ($M)"] = (icp_list["Assets ($M)"] / 1e6).round(0).astype(int)
    icp_list["Inv CRE ($M)"] = (icp_list["Inv CRE ($M)"] / 1e6).round(0).astype(int)
    icp_list["Investor CRE %"] = icp_list["Investor CRE %"].round(0).astype(int)
    icp_list["C&I %"] = icp_list["C&I %"].round(0).astype(int)
    table(icp_list.reset_index(drop=True), (
        "Ready-made outreach list for investor CRE borrowers targeting ICP community banks. "
        "Banks are ranked by investor CRE portfolio share — the top names have made income-property lending "
        "a core competency. Check C&I % for banks that can also provide operating lines alongside property debt. "
        "Cross-reference City with your asset location; HQ city does not guarantee local lending presence."
    ))

    insight(
        id="I-06", priority="High", theme="ICP opportunity", section="EDA §8",
        insight=(
            f"{len(icp)} banks in $500M–$2B band. ICP specialists: C&I {icp_spec.get('Commercial & Industrial (C&I)', 0)}, "
            f"Investor CRE {icp_spec.get('Investor CRE (income property)', 0)}, "
            f"Multifamily {icp_spec.get('Multifamily (5+ units)', 0)}."
        ),
        explanation=(
            "The ICP band balances capacity ($2M–$25M relationship credits) with accessibility (direct sponsor relationships). "
            "These banks are large enough to hold meaningful commercial credits but small enough to value borrower relationships."
        ),
        evidence="total_assets from texas_bank_profiles_latest.csv.",
        borrower_action="Shortlist 3–5 ICP banks in your city with ≥8% in your product before broad outreach.",
    )

    # ── 9. Geography ────────────────────────────────────────────────────────
    section("9. Geographic lender density (HQ city, top 25)")
    if active["city"].notna().any():
        city_stats = (
            active.groupby("city", dropna=False)
            .agg(
                banks=("id_rssd", "count"),
                median_loans_M=("total_loans_usd", lambda s: int(round(s.median() / 1e6))),
                inv_spec=("mix_inv_pct", lambda s: int((s >= SPECIALIST_PCT).sum())),
                ci_spec=("mix_ci_pct", lambda s: int((s >= SPECIALIST_PCT).sum())),
                mf_spec=("mix_mf_pct", lambda s: int((s >= SPECIALIST_PCT).sum())),
            )
            .reset_index()
            .sort_values("banks", ascending=False)
            .head(25)
        )
        city_stats.columns = ["City", "Banks (HQ)", "Median loans ($M)", "Inv CRE spec.", "C&I spec.", "MF spec."]
        table(city_stats, (
            "Dallas and Houston lead in bank HQ concentration and specialist depth. More bank HQs generally means "
            "more lender choice for borrowers in those metros, but also more competition for banker attention. "
            "Specialist columns count banks with ≥8% portfolio share in each product — use this to prioritize "
            "cities when building a geographic outreach strategy. Remember: HQ city is not the same as branch "
            "footprint; a bank headquartered in Dallas may lend statewide."
        ))

        dallas = city_stats[city_stats["City"].str.upper() == "DALLAS"]
        if not dallas.empty:
            d = dallas.iloc[0]
            insight(
                id="I-07", priority="Medium", theme="Geography", section="EDA §9",
                insight=(
                    f"Dallas: {int(d['Banks (HQ)'])} bank HQs, {int(d['Inv CRE spec.'])} investor CRE specialists, "
                    f"{int(d['C&I spec.'])} C&I specialists."
                ),
                explanation="HQ city density is a proxy for lender choice. Major metros offer more options but require sharper differentiation.",
                evidence="HQ city from bank profiles; does not reflect branch footprint.",
                borrower_action="Verify branch presence — HQ city ≠ lending territory.",
            )

    # ── 10. Quarterly trends ────────────────────────────────────────────────
    section("10. Texas loan market trends by quarter (aggregate)")
    trend_rows = []
    for period_dt, grp in df.groupby("period_dt"):
        m = enrich_mix(compute_mix_frame(build_bank_wide(grp, profiles)), profiles)
        sub = m[m["total_loans_gross"] > 0]
        trend_rows.append({
            "Period": period_dt.strftime("%Y-%m-%d"),
            "Banks": len(sub),
            "Total loans ($B)": int(round(sub["total_loans_usd"].sum() / 1e9)),
            "Investor CRE ($B)": int(round(sub["mix_inv_usd"].sum() / 1e9)),
            "C&I ($B)": int(round(sub["mix_ci_usd"].sum() / 1e9)),
            "Multifamily ($B)": int(round(sub["mix_mf_usd"].sum() / 1e9)),
            "Construction ($B)": int(round(sub["mix_con_usd"].sum() / 1e9)),
        })
    trend_df = pd.DataFrame(trend_rows).sort_values("Period")
    table(trend_df, (
        "This shows how aggregate Texas bank loan books evolved over five quarters. Rising totals in your product "
        "sector suggest growing lender capacity and appetite; flat or declining lines may signal tighter underwriting "
        "or portfolio runoff. The bank count column also shifts slightly each quarter as institutions file or "
        "merge. Use this to time outreach — approaching lenders when their sector book is growing often yields "
        "better reception than during contraction periods."
    ))

    section("10b. Quarterly trends by borrower product ($B)")
    prod_trend = []
    for period_dt, grp in df.groupby("period_dt"):
        m = compute_mix_frame(build_bank_wide(grp, profiles))
        sub = m[m["total_loans_gross"] > 0]
        row = {"Period": period_dt.strftime("%Y-%m-%d")}
        for key, (label, _) in MIX_DEF.items():
            row[label[:28]] = int(round(sub[f"mix_{key}_usd"].sum() / 1e9))
        prod_trend.append(row)
    table(pd.DataFrame(prod_trend).sort_values("Period"), (
        "Product-level quarterly breakdown — each column is the aggregate statewide balance ($B) for that borrower "
        "category. Compare your product's row across periods to spot growth or contraction. Investor CRE and C&I "
        "typically dominate Texas bank books; multifamily and ag are smaller but may be growing in specific quarters. "
        "Pair this with the specialist tables to identify which banks are driving sector growth."
    ))

    if len(trend_df) >= 2:
        first, last = trend_df.iloc[0], trend_df.iloc[-1]
        chg = (last["Total loans ($B)"] - first["Total loans ($B)"]) / first["Total loans ($B)"] * 100
        insight(
            id="I-08", priority="Medium", theme="Market momentum", section="EDA §10",
            insight=f"Total loans: ${first['Total loans ($B)']}B → ${last['Total loans ($B)']}B ({chg:+.1f}%) over {len(trend_df)} quarters.",
            explanation=(
                "Quarter-over-quarter portfolio growth can signal lender appetite for new originations. "
                "Flat or declining books in your sector may mean tighter underwriting."
            ),
            evidence="Five-quarter aggregate from texas_loans_summary.csv.",
            borrower_action="Ask lenders if their portfolio in your asset class grew last quarter.",
        )

    # ── 11. QoQ growth ──────────────────────────────────────────────────────
    section("11. Quarter-over-quarter growth by product (latest vs prior quarter)")
    periods = sorted(df["period_dt"].unique())
    if len(periods) >= 2:
        prev, curr = periods[-2], periods[-1]
        m_prev = compute_mix_frame(build_bank_wide(df[df["period_dt"] == prev], profiles))
        m_curr = compute_mix_frame(build_bank_wide(df[df["period_dt"] == curr], profiles))
        qoq = []
        for key, (label, _) in MIX_DEF.items():
            p = m_prev[f"mix_{key}_usd"].sum()
            c = m_curr[f"mix_{key}_usd"].sum()
            qoq.append({
                "Product": label,
                f"Prior ({prev.strftime('%Y-%m-%d')}) $B": int(round(p / 1e9)),
                f"Latest ({curr.strftime('%Y-%m-%d')}) $B": int(round(c / 1e9)),
                "QoQ change %": int(round(100 * (c - p) / p)) if p > 0 else None,
            })
        table(pd.DataFrame(qoq), (
            f"Compares the most recent quarter ({curr.strftime('%Y-%m-%d')}) to the prior quarter ({prev.strftime('%Y-%m-%d')}) "
            f"for each product category. Positive QoQ change % means Texas banks collectively grew that book — "
            f"a signal of lender appetite. Negative change may reflect paydowns, sales, or tighter origination. "
            f"Focus outreach on products and banks showing positive momentum in your sector."
        ))

    # ── 12. Lender health ─────────────────────────────────────────────────────
    section("12. Portfolio stress — 90+ day past-due (% of total loans)")
    stress = active.copy()
    stress["stress_flag"] = stress["past_due_90_plus_pct"] > 1.0
    table(pd.DataFrame([
        {"Metric": "Median 90+ day past-due %", "Value": int(round(stress["past_due_90_plus_pct"].median()))},
        {"Metric": "75th percentile", "Value": int(round(stress["past_due_90_plus_pct"].quantile(0.75)))},
        {"Metric": "Banks with elevated stress (>1%)", "Value": int(stress["stress_flag"].sum())},
        {"Metric": "% of banks elevated", "Value": int(round(100 * stress["stress_flag"].mean()))},
    ]), (
        "Portfolio-level credit health across Texas banks. The median 90+ day past-due ratio is very low for most "
        "banks, meaning the overall market looks healthy. Banks above 1% deserve extra diligence — ask whether "
        "they are still originating in your asset class or conserving capital. This is not a reason to avoid a "
        "bank automatically, but it is a conversation starter about recent charge-offs and underwriting posture."
    ))

    section("12b. Banks with highest 90+ day past-due ratios (diligence flags)")
    stress_top = stress.nlargest(15, "past_due_90_plus_pct")[
        ["institution_name", "city", "past_due_90_plus_pct", "total_loans_usd", "mix_inv_pct", "mix_ci_pct"]
    ].copy()
    stress_top.columns = ["Bank", "City", "90+ past-due %", "Total loans ($M)", "Inv CRE %", "C&I %"]
    stress_top["90+ past-due %"] = stress_top["90+ past-due %"].round(0).astype(int)
    stress_top["Total loans ($M)"] = (stress_top["Total loans ($M)"] / 1e6).round(0).astype(int)
    stress_top["Inv CRE %"] = stress_top["Inv CRE %"].round(0).astype(int)
    stress_top["C&I %"] = stress_top["C&I %"].round(0).astype(int)
    table(stress_top.reset_index(drop=True), (
        "These fifteen banks have the highest 90+ day past-due ratios relative to total loans. "
        "Borrowers should not automatically exclude them, but should ask direct questions: Are you still booking "
        "new loans in my sector? What drove recent delinquencies? Is there a concentration issue? "
        "Banks with high past-due but strong CRE or C&I mix may still be active — the ratio reflects legacy issues, "
        "not necessarily current appetite."
    ))

    insight(
        id="I-09", priority="Medium", theme="Lender diligence", section="EDA §12",
        insight=(
            f"{int(stress['stress_flag'].sum())} banks ({100 * stress['stress_flag'].mean():.1f}%) have 90+ day past-due >1%. "
            f"Market median: {stress['past_due_90_plus_pct'].median():.2f}%."
        ),
        explanation=(
            "Past-due ratios are a portfolio-level health signal. Elevated stress does not mean a bank won't lend — "
            "but it warrants questions about credit appetite, charge-offs, and whether new originations are paused."
        ),
        evidence="past_due_90_plus from texas_bank_profiles_latest.csv vs RCON2122 total loans.",
        borrower_action="Ask about recent charge-offs and appetite for new commitments in your asset class.",
    )

    # ── 13. Archetypes ──────────────────────────────────────────────────────
    section("13. Portfolio archetypes — dominant product per bank")
    archetypes = []
    for _, r in active.iterrows():
        mixes = {k: r[f"mix_{k}_pct"] for k in MIX_DEF}
        top_k = max(mixes, key=mixes.get)
        archetypes.append({
            "dominant": MIX_DEF[top_k][0],
            "dom_pct": int(round(mixes[top_k])),
            "cre_ci": int(round(r["mix_inv_pct"] + r["mix_own_pct"] + r["mix_ci_pct"])),
            "consumer": int(round(r["mix_cons_pct"])),
        })
    arch_df = pd.DataFrame(archetypes)
    arch_sum = (
        arch_df.groupby("dominant")
        .agg(banks=("dom_pct", "count"), median_dom=("dom_pct", "median"))
        .reset_index()
        .sort_values("banks", ascending=False)
    )
    arch_sum.columns = ["Dominant product", "Banks", "Median dominance %"]
    arch_sum["Median dominance %"] = arch_sum["Median dominance %"].round(0).astype(int)
    table(arch_sum, (
        "Each bank is classified by its single largest loan category. The 'Banks' column shows how many Texas lenders "
        "are primarily oriented toward that product — for example, banks dominated by owner-occupied CRE vs C&I. "
        "Commercial borrowers should seek banks whose dominant product matches their deal type, or banks classified "
        "as portfolio-style (CRE + C&I combined > 50%) in section 7. Consumer-dominated banks are weak fits for CRE outreach."
    ))

    port_n = int((arch_df["cre_ci"] > 50).sum())
    cons_n = int((arch_df["consumer"] > 25).sum())
    insight(
        id="I-10", priority="High", theme="Bank archetypes", section="EDA §13",
        insight=(
            f"{port_n} portfolio-style banks (CRE+C&I >50%); {cons_n} consumer-heavy (>25%). "
            f"Most common dominant type: {arch_sum.iloc[0]['Dominant product']} ({int(arch_sum.iloc[0]['Banks'])} banks)."
        ),
        explanation=(
            "Archetyping groups banks by what they mostly do. Commercial borrowers should filter OUT consumer-heavy banks "
            "and prioritize portfolio-style lenders where CRE and C&I are the core business."
        ),
        evidence="Dominant product = highest mix % among borrower categories.",
        borrower_action="Filter consumer-heavy banks before CRE/C&I outreach.",
    )

    # ── 14. Co-occurrence ─────────────────────────────────────────────────────
    section("14. Product co-occurrence among specialist lenders (%)")
    keys = list(MIX_DEF.keys())[:8]
    spec_matrix = []
    for k in keys:
        row = {"Row product": MIX_DEF[k][0][:30]}
        base = active[f"mix_{k}_pct"] >= SPECIALIST_PCT
        for k2 in keys:
            if k == k2:
                row[MIX_DEF[k2][0][:18]] = 100.0
            else:
                both = (base & (active[f"mix_{k2}_pct"] >= SPECIALIST_PCT)).sum()
                row[MIX_DEF[k2][0][:18]] = int(round(100 * both / base.sum())) if base.sum() else 0
        spec_matrix.append(row)
    table(pd.DataFrame(spec_matrix), (
        "Each cell shows what percentage of row-product specialists are also column-product specialists. "
        "Diagonal values are 100% by definition. High off-diagonal values (e.g. Investor CRE row × C&I column) "
        "mean banks that do one product seriously often do the other — useful if you need property debt plus an "
        "operating line from the same relationship. Low overlap (e.g. multifamily × ag) means you need separate "
        "lenders for each product."
    ))

    insight(
        id="I-11", priority="Low", theme="Cross-sell patterns", section="EDA §14",
        insight="Investor CRE specialists frequently overlap with C&I specialists — full-relationship banks can bundle operating lines with property loans.",
        explanation=(
            "Co-occurrence shows what % of specialists in one product are also specialists in another. "
            "High overlap means one bank can potentially handle multiple parts of your capital stack."
        ),
        evidence="Co-specialization matrix among banks with ≥8% in each product.",
        borrower_action="If you need CRE plus a revolver, prioritize banks with high co-occurrence.",
    )

    # ── 15. MDRM code inventory ───────────────────────────────────────────────
    section("15. MDRM code inventory in this summary file")
    code_stats = (
        latest.groupby(["mdrm_code", "item_name"])
        .agg(
            banks=("id_rssd", "nunique"),
            nonzero=("value_num", lambda s: int((s > 0).sum())),
            total_B=("value_num", lambda s: int(round(s.sum() / 1e9))),
            median_M=("value_num", lambda s: int(round(s[s > 0].median() / 1e6)) if (s > 0).any() else 0),
        )
        .reset_index()
        .sort_values("nonzero", ascending=False)
    )
    code_stats.columns = ["Code", "Fed label (MDRM)", "Banks", "Nonzero", "TX total ($B)", "Median nonzero ($M)"]
    table(code_stats, (
        "Inventory of every regulatory line in this summary file. 'Nonzero' counts how many bank-quarters report a "
        "balance > 0 — high counts mean broad market participation; low counts mean niche lines. "
        "RCONF161 (investor CRE) and RCON2122 (total loans) have near-universal coverage. "
        "Sparse lines like RCONF162 or RCON1545 reflect form-type differences (041 vs 031) or product rarity in Texas."
    ))

    section("15b. Codes with sparse Texas coverage (latest quarter)")
    sparse = code_stats[code_stats["Nonzero"] < 50][["Code", "Fed label (MDRM)", "Nonzero", "Banks"]]
    if not sparse.empty:
        table(sparse, (
            "These MDRM codes appear in fewer than 50 bank-quarters with non-zero balances. "
            "Borrowers should not rely solely on these lines for lender identification — supplement with "
            "related codes (e.g. use RCONF161 for investor CRE when RCONF162 is sparse) or the full "
            "texas_loans_labeled.csv for deeper RC-C detail."
        ))

    # ── 16. Portfolio size buckets ────────────────────────────────────────────
    section("16. Bank count by total loan portfolio size")
    buckets = pd.cut(
        active["total_loans_usd"],
        bins=[0, 50e6, 250e6, 1e9, 5e9, 50e9, np.inf],
        labels=["<$50M", "$50–250M", "$250M–$1B", "$1–5B", "$5–50B", ">$50B"],
    )
    bucket_df = (
        active.groupby(buckets, observed=True)
        .agg(banks=("id_rssd", "count"), median_inv_pct=("mix_inv_pct", "median"), median_ci_pct=("mix_ci_pct", "median"))
        .reset_index()
    )
    bucket_df.columns = ["Portfolio size", "Banks", "Median inv CRE %", "Median C&I %"]
    bucket_df["Median inv CRE %"] = bucket_df["Median inv CRE %"].round(0).astype(int)
    bucket_df["Median C&I %"] = bucket_df["Median C&I %"].round(0).astype(int)
    table(bucket_df, (
        "Distribution of Texas banks by total loan portfolio size. Most community banks sit below $1B in total loans. "
        "Median mix columns show how CRE and C&I orientation shifts by size band — mid-size banks ($250M–$5B) often "
        "show the strongest commercial mix. Match your deal size to a bank's size bucket: a $5M CRE loan fits "
        "naturally in the $250M–$1B band; a $50M ticket requires banks in the $5B+ bands."
    ))

    # ── 17. Specialist shortlist sizing ───────────────────────────────────────
    section("17. Recommended shortlist size by product")
    table(pd.DataFrame([
        {
            "Product": MIX_DEF[k][0],
            f"≥{int(SPECIALIST_PCT)}%": int((active[f"mix_{k}_pct"] >= SPECIALIST_PCT).sum()),
            "≥15%": int((active[f"mix_{k}_pct"] >= 15).sum()),
            "Suggested calls": min(15, int((active[f"mix_{k}_pct"] >= SPECIALIST_PCT).sum())),
        }
        for k in MIX_DEF
    ]), (
        "Practical outreach planning guide. 'Suggested calls' caps at 15 because beyond that, marginal returns drop. "
        "C&I and Investor CRE offer the deepest specialist pools — you can run a full competitive process. "
        "Multifamily and Construction have far fewer specialists, so each call must be well-researched. "
        "Use ≥8% as your first filter, then narrow by geography, ICP fit, and deal size."
    ))

    insight(
        id="I-12", priority="High", theme="Outreach planning", section="EDA §17",
        insight="Build a 10–15 bank shortlist for C&I and Investor CRE; Multifamily has only ~19 deep specialists statewide.",
        explanation=(
            "The 'suggested calls' column caps at 15 — beyond that, returns diminish. For niche products, "
            "quality of fit beats quantity of outreach."
        ),
        evidence="Specialist counts at ≥8% and ≥15% portfolio share thresholds.",
        borrower_action="Prepare a one-page deal summary; lead with portfolio fit.",
    )

    # ── 18. Mix percentiles ───────────────────────────────────────────────────
    section("18. Portfolio mix percentiles by product (active lenders, latest)")
    pct_rows = []
    for key, (label, _) in MIX_DEF.items():
        col = f"mix_{key}_pct"
        pos = active.loc[active[col] > 0, col]
        if pos.empty:
            continue
        pct_rows.append({
            "Product": label,
            "P25 %": int(round(pos.quantile(0.25))),
            "Median %": int(round(pos.median())),
            "P75 %": int(round(pos.quantile(0.75))),
            "Max %": int(round(pos.max())),
        })
    table(pd.DataFrame(pct_rows), (
        "Shows how concentrated each product is within banks that hold any balance in it. "
        "If P75 is only 5% but you need a specialist, you are looking for banks above the 75th percentile — "
        "typically the ≥8% threshold used elsewhere in this analysis. High max values (e.g. 80%+) indicate "
        "banks that are essentially single-product lenders — powerful allies for that niche but poor fits for "
        "other product types."
    ))

    # ── 19. Taxonomy crosswalk ────────────────────────────────────────────────
    if TAXONOMY.exists():
        section("19. Taxonomy crosswalk (texas_mdrm_loan_taxonomy.csv)")
        tax = pd.read_csv(TAXONOMY, low_memory=False)
        if "Loan Product Category" in tax.columns:
            cov = (
                tax.groupby("Loan Product Category")
                .agg(
                    regulatory_lines=("Regulatory Line Item Code", "nunique"),
                    median_banks=("Number of Texas Banks Reporting", "median"),
                )
                .reset_index()
                .sort_values("median_banks", ascending=False)
                .head(15)
            )
            table(cov, (
                "Cross-reference to the full texas_mdrm_loan_taxonomy dataset. Shows how many distinct regulatory "
                "lines map to each borrower-facing product category and the typical number of Texas banks reporting "
                "each category. Use this to understand which products have rich regulatory granularity (many lines) "
                "versus those captured by a single MDRM code."
            ))

    # ── Methodology ───────────────────────────────────────────────────────────
    section("Methodology, assumptions & limitations")
    eda.extend([
        ("Data source", str(INPUT.relative_to(REPO))),
        ("Grain", "One row = bank × quarter × MDRM regulatory line item"),
        ("Borrower mapping", "Schedule RC-C summary codes → Lenni loan product taxonomy"),
        ("Specialist definition", f"Product balance ≥ {SPECIALIST_PCT}% of total loans (RCON2122)"),
        ("C&I source", "RCON1766 from texas_bank_profiles_latest.csv (full XBRL pivot)"),
        ("Units", "value_num and profile fields are USD"),
        ("Limitation", "Regulatory categories ≠ bank marketing product names"),
        ("Limitation", "No pricing, LTV, credit box, or committee thresholds"),
        ("Limitation", "HQ city only — branch footprint requires separate data"),
        ("Limitation", "Past-due is portfolio-level, not product-specific"),
        ("Limitation", "RCON5367/5368 are residential lien balances, not past-due amounts"),
        ("", ""),
    ])

    return eda, insights


def build_provenance_rows(df: pd.DataFrame) -> list[tuple[str, str]]:
    """Narrative documentation of the full API → CSV → XLSX pipeline."""
    rows: list[tuple[str, str]] = []
    n_progress = 0
    if PROGRESS.is_file():
        n_progress = len(json.loads(PROGRESS.read_text()).get("completed", []))

    n_filings = n_inst = 0
    period_counts = ""
    if FILINGS.is_file():
        fil = pd.read_csv(FILINGS)
        n_filings = len(fil)
        if "reporting_period" in fil.columns:
            pc = fil["reporting_period"].value_counts().sort_index()
            period_counts = "; ".join(f"{p}: {c} filings" for p, c in pc.items())

    if INSTITUTIONS.is_file():
        n_inst = len(pd.read_csv(INSTITUTIONS))

    periods = sorted(pd.to_datetime(df["reporting_period"], dayfirst=False).unique())
    period_list = ", ".join(pd.Timestamp(p).strftime("%Y-%m-%d") for p in periods)

    def sec(title: str) -> None:
        rows.append((f"§ {title}", ""))
        rows.append(("", ""))

    rows.append(("DATA PROVENANCE — How texas_loans_summary was built", ""))
    rows.append(("", ""))
    rows.extend([
        ("Document built", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Output workbook", "texas_loans_summary.xlsx (this file, sheet 1)"),
        ("Source CSV", "ONLY_TEXAS_SINCE_2025/exports/texas_loans_summary.csv"),
        ("Project", "ONLY_TEXAS_SINCE_2025 — Texas Call Reports (2025+)"),
        ("Maintainer context", "PayPro LLC / FFIEC_DATABASE_PROJECT"),
        ("", ""),
    ])

    sec("1. Executive summary")
    rows.extend([
        ("What this data is", (
            "Aggregated loan portfolio balances reported by Texas banks to federal regulators "
            "(FFIEC Call Report, Schedule RC-C and related lines). Each row is one regulatory "
            "category per bank per quarter — not individual loans."
        )),
        ("Geographic scope", "Texas only (State = TX on FFIEC Panel of Reporters)"),
        ("Time scope", f"Reporting periods ending 2025 or later: {period_list}"),
        ("Rows in this sheet", f"{len(df):,}"),
        ("Banks covered", f"{df['id_rssd'].nunique()} distinct RSSD IDs across all periods"),
        ("MDRM codes in summary", f"{df['mdrm_code'].nunique()} curated Schedule RC-C lines"),
        ("Authentication", "No web scraping — official FFIEC Public Web Service (REST API) + Federal Reserve MDRM dictionary"),
        ("", ""),
    ])

    sec("2. Phase A — FFIEC API download (pull_texas_since_2025.py)")
    rows.extend([
        ("Script", "ONLY_TEXAS_SINCE_2025/pull_texas_since_2025.py"),
        ("API base URL", "https://ffieccdr.azure-api.us/public/"),
        ("API specification", "https://cdr.ffiec.gov/public/Files/SIS611_-_Retrieve_Public_Data_via_Web_Service.pdf"),
        ("Credentials", "FFIEC_USER_ID + FFIEC_TOKEN in ffiec-cdr/.env (Bearer token auth)"),
        ("Rate limiting", "~1.5 seconds between API calls (~2,400/hour, under FFIEC 2,500/hour cap)"),
        ("", ""),
        ("Step 0 — Authenticate", "HTTP headers: UserID + Authentication: Bearer <token>"),
        ("Step 1 — List periods", "API: RetrieveReportingPeriods (data_series=Call) → filter year ≥ 2025"),
        ("Step 2 — Panel of reporters", "API: RetrievePanelOfReporters per period → filter State == TX"),
        ("Step 3 — Write institutions", "All TX panel rows → exports/texas_institutions.csv"),
        ("Step 4 — Download filings", "For each bank with HasFiledForReportingPeriod=True: RetrieveFacsimile (XBRL format)"),
        ("Step 5 — Archive raw files", "Save to archive/call/<period>/<rssd>.xbrl + .meta.json (SHA-256, timestamp)"),
        ("Step 6 — Parse XBRL", "ffiec_cdr.parser.parse_xbrl → append exports/texas_xbrl_facts.csv"),
        ("Step 7 — Resume support", "data/progress.json tracks completed period|RSSD pairs — skips on re-run"),
        ("", ""),
        ("Download results (checkpoint)", f"{n_progress:,} period|RSSD pairs completed in progress.json"),
        ("Filings inventory rows", f"{n_filings:,} in texas_filings.csv"),
        ("Institution panel rows", f"{n_inst:,} in texas_institutions.csv"),
        ("Filings by period", period_counts or "See texas_filings.csv"),
        ("Raw archive", "1,825 .xbrl files under ONLY_TEXAS_SINCE_2025/archive/call/"),
        ("Parsed fact rows", "~2,186,590 rows in texas_xbrl_facts.csv (all line items, all banks)"),
        ("Runtime", "Typically 1–3 hours for full Texas 2025+ pull"),
        ("", ""),
    ])

    sec("3. FFIEC APIs used (and not used)")
    rows.extend([
        ("RetrieveReportingPeriods", "YES — list Call Report quarters"),
        ("RetrievePanelOfReporters", "YES — bank panel with state, city, filing status"),
        ("RetrieveFacsimile", "YES — download XBRL Call Report per bank/quarter"),
        ("RetrieveFilersSinceDate", "NO — incremental sync only (not needed for full backfill)"),
        ("RetrieveFilersSubmissionDateTime", "NO — optional submission timestamps"),
        ("RetrieveUBPRReportingPeriods", "NO — UBPR is a different report (peer ratios)"),
        ("RetrieveUBPRXBRLFacsimile", "NO — UBPR XBRL, not Call Report"),
        ("Formats not downloaded", "PDF and SDF (same API with different facsimileFormat header)"),
        ("", ""),
    ])

    sec("4. Phase B — MDRM labeling (extract_texas_loans.py --summary)")
    rows.extend([
        ("Script", "ONLY_TEXAS_SINCE_2025/extract_texas_loans.py --summary"),
        ("MDRM dictionary", "https://www.federalreserve.gov/apps/mdrm/pdf/MDRM.zip → data/mdrm/MDRM_CSV.csv"),
        ("MDRM loader", "ONLY_TEXAS_SINCE_2025/mdrm_loader.py"),
        ("Input", "exports/texas_xbrl_facts.csv (~2.19M rows)"),
        ("Filter logic", "Keep curated SUMMARY_CODES (Schedule RC-C totals, CRE, consumer, ag, construction, etc.)"),
        ("Enrichment", "Join each mdrm_code to Fed item_name, mdrm_description, mdrm_category, reporting_form"),
        ("Output", "exports/texas_loans_summary.csv — 31,396 rows, 21 MDRM codes"),
        ("Also produced", "texas_loans_labeled.csv (~938k rows, full RC-C) and texas_loan_products_mdrm_catalog.csv"),
        ("", ""),
        ("Summary codes included", (
            "RCON2122/RCFD2122 (total loans), RCONF161/160/158/159 (CRE categories), RCON1460 (multifamily), "
            "RCON1403 (1-4 family), RCON1545/1583 (consumer), RCON1420/1590 (ag), RCON1754 (lease), "
            "RCON1797 (HELOC), RCON5367/5368 (residential liens), RCON2130/2145/5369, etc."
        )),
        ("Codes in summary but sparse in TX", "RCON1480, RCON1766 (C&I) — supplemented in EDA from bank profiles"),
        ("", ""),
    ])

    sec("5. Phase C — Supporting joins used in EDA (this workbook)")
    rows.extend([
        ("texas_bank_profiles_latest.csv", "Latest-quarter wide metrics per bank: total_assets, ci_loans (RCON1766), past_due_90_plus, icp_fit"),
        ("Join key", "id_rssd (Federal Reserve RSSD ID)"),
        ("ICP definition", "$500M–$2B total assets (community bank sweet spot for commercial borrowers)"),
        ("build_lenni_eda_report.py", "Optional: joins all CSVs and generates 25-chart PDF report"),
        ("", ""),
    ])

    sec("6. Data model — how tables relate")
    rows.extend([
        ("texas_institutions", "Grain: bank × quarter. Panel of reporters (who must file)."),
        ("texas_filings", "Grain: bank × quarter. Download inventory with SHA-256 and file path."),
        ("texas_xbrl_facts", "Grain: one row per XBRL fact (concept × context). Source of all numeric values."),
        ("texas_loans_summary", "Grain: bank × quarter × MDRM loan line. Filtered + labeled subset for analysis."),
        ("Join keys", "id_rssd + reporting_period links all tables"),
        ("", ""),
    ])

    sec("7. Rebuild & recovery commands")
    rows.extend([
        ("Full re-download", "python ONLY_TEXAS_SINCE_2025/pull_texas_since_2025.py"),
        ("Test download (5 filings)", "python ONLY_TEXAS_SINCE_2025/pull_texas_since_2025.py --max 5"),
        ("Rebuild CSVs from archive", "python ONLY_TEXAS_SINCE_2025/rebuild_csv_from_archive.py"),
        ("Regenerate loan summary", "python ONLY_TEXAS_SINCE_2025/extract_texas_loans.py --summary"),
        ("Download MDRM dictionary", "python ONLY_TEXAS_SINCE_2025/scripts/download_mdrm.py"),
        ("Rebuild this workbook", "python ONLY_TEXAS_SINCE_2025/build_texas_loans_summary_borrower_eda.py"),
        ("", ""),
    ])

    sec("8. What this data is NOT")
    rows.extend([
        ("Individual loans", "Call Reports are confidential aggregates — no borrower names or loan-level detail"),
        ("Marketing product names", "Only regulatory categories (e.g. 'other nonfarm nonresidential') — not '30-year fixed'"),
        ("Pricing or terms", "No interest rates, LTV limits, or credit policies"),
        ("Real-time data", "Quarterly regulatory filings with reporting lag"),
        ("National scope", "Texas banks only — not all US institutions"),
        ("UBPR peer ratios", "Requires separate UBPR API pull (not included)"),
        ("Branch geography", "HQ city from panel — branch footprint requires separate locations data"),
        ("", ""),
    ])

    sec("9. References & documentation")
    rows.extend([
        ("FFIEC CDR public site", "https://cdr.ffiec.gov/public/"),
        ("PWS help", "https://cdr.ffiec.gov/public/HelpFiles/PWSInfo.htm"),
        ("Federal Reserve MDRM", "https://www.federalreserve.gov/apps/mdrm/"),
        ("MDRM code lookup", "https://www.federalreserve.gov/apps/mdrm/data-dictionary"),
        ("Schedule RC-C instructions", "https://www.fdic.gov/bank-financial-reports/031-041-rc-c1-loans-and-leases-december-2024"),
        ("Project README", "ONLY_TEXAS_SINCE_2025/README.md"),
        ("Data dictionary", "ONLY_TEXAS_SINCE_2025/DATA_DICTIONARY.md"),
        ("Loan extraction guide", "ONLY_TEXAS_SINCE_2025/LOAN_EXTRACTION_GUIDE.md"),
        ("SharePoint data folder", "See README.md for teammate SharePoint link"),
        ("", ""),
    ])

    sec("10. Project milestones")
    rows.extend([
        ("2026-06-02", "Initial Texas extract script; first XBRL downloads"),
        ("2026-06-02–03", "Full backfill: 1,825 filings across 5 quarters"),
        ("2026-06-03", "rebuild_csv_from_archive.py; CSVs aligned with archive"),
        ("2026-06-03", "Federal Reserve MDRM integration; loan-labeled exports"),
        ("2026-06-03", "SharePoint distribution for teammates"),
        ("", ""),
    ])

    return rows


# Borrower-friendly labels when Fed MDRM item_name in the extract does not match RC-C intent.
MDRM_BORROWER_LABEL = {
    "RCON2122": "Total loans & leases, net of unearned income (headline portfolio size)",
    "RCFD2122": "Total loans & leases — domestic/consolidated variant (smaller-bank form)",
    "RCON2145": "Net loans (intended RC-C line; verify against RCON2122 in source XBRL)",
    "RCON2130": "Allowance for loan losses (intended line; label may vary by taxonomy quarter)",
    "RCON1403": "1–4 family residential mortgage loans (Schedule RC-C)",
    "RCON1754": "Lease financing receivables",
    "RCON1545": "Credit card plans",
    "RCON1583": "Other consumer loans",
    "RCON5367": "1–4 family residential loans secured by first lien",
    "RCON5368": "1–4 family residential loans secured by junior lien",
    "RCON1766": "Commercial & industrial (C&I) loans — from bank profiles, not in summary CSV",
    "RCON1480": "Commercial & industrial loans (alternate MDRM line — sparse in Texas summary)",
}


def build_abbreviations_glossary(df: pd.DataFrame) -> pd.DataFrame:
    """Every abbreviation, code, and acronym used in the dataset and EDA workbook."""
    rows: list[dict] = []

    def add(section: str, abbr: str, full_term: str, definition: str) -> None:
        rows.append({
            "Section": section,
            "Abbreviation / code": abbr,
            "Full term": full_term,
            "Definition & usage in this workbook": definition,
        })

    # ── A. MDRM regulatory line codes in texas_loans_summary ─────────────────
    code_meta = (
        df.groupby("mdrm_code")
        .agg(
            item_name=("item_name", "first"),
            mdrm_category=("mdrm_category", "first"),
            banks_reporting=("id_rssd", "nunique"),
        )
        .reset_index()
        .sort_values("mdrm_code")
    )
    for _, r in code_meta.iterrows():
        code = r["mdrm_code"]
        borrower = MDRM_BORROWER_LABEL.get(code, "")
        fed_label = str(r["item_name"]).strip()
        definition = (
            f"Federal Reserve MDRM line code on Schedule RC-C / Call Report. "
            f"Category: {r['mdrm_category']}. "
            f"Reported by {int(r['banks_reporting'])} Texas banks in this file. "
            f"Fed dictionary label: {fed_label}."
        )
        if borrower:
            definition += f" Borrower interpretation: {borrower}."
        add("A — MDRM codes in dataset", code, fed_label, definition)

    # Related codes referenced in EDA but not always in summary rows
    for code, borrower in sorted(MDRM_BORROWER_LABEL.items()):
        if code not in code_meta["mdrm_code"].values:
            add(
                "A — MDRM codes in dataset",
                code,
                borrower.split("—")[0].strip(),
                f"Referenced in EDA joins (e.g. bank profiles). {borrower}.",
            )

    # ── B. MDRM code prefix conventions ───────────────────────────────────────
    prefixes = [
        ("RCON", "Report Code — domestic office", "Dollar amounts for domestic offices of the reporting bank; most common prefix for large banks."),
        ("RCFD", "Report Code — consolidated/domestic", "Same economic line as RCON#### but for consolidated or domestic-office reporting (often FFIEC 041 filers)."),
        ("RCONF", "RC-C detail extension", "Schedule RC-C sub-lines for loan categories (construction, CRE, leases)."),
        ("RCON14xx", "Real estate secured loans", "Prefix family for mortgage and CRE collateral types."),
        ("RCON15xx", "Consumer loans", "Credit cards, other consumer, revolving credit."),
        ("RCON16xx", "Other loans", "C&I and other non-RE loan categories."),
        ("RCON21xx", "Loan totals & allowance", "Headline total loans, allowance, net loans."),
        ("RCON53xx", "Residential mortgage detail", "First lien, junior lien, and related 1–4 family breakdowns."),
    ]
    for abbr, full, defn in prefixes:
        add("B — MDRM prefix families", abbr, full, defn)

    # ── C. Source file column names ────────────────────────────────────────────
    columns = [
        ("id_rssd", "RSSD identification number", "Federal Reserve unique ID for the institution; primary join key."),
        ("institution_name", "Institution name", "Legal or reporting name from the FFIEC panel of reporters."),
        ("reporting_period", "Reporting period", "Call Report quarter-end date (MM/DD/YY or MM/DD/YYYY)."),
        ("mdrm_code", "MDRM code", "Micro Data Reference Manual line identifier (e.g. RCON2122)."),
        ("item_name", "Item name", "Short official label from the Federal Reserve MDRM dictionary."),
        ("line_description", "Line description", "Duplicate short label for spreadsheet compatibility."),
        ("mdrm_description", "MDRM description", "Full Federal Reserve definition of the regulatory line (up to ~800 characters)."),
        ("mdrm_category", "MDRM category", "loan_or_lease, schedule_rc_c, or loan_related_prefix — how the line is classified."),
        ("reporting_form", "Reporting form", "FFIEC Call Report form filed (031, 041, etc.)."),
        ("item_type", "Item type", "F = financial line item; D = derived or memorandum item."),
        ("value_num", "Numeric value", "Reported balance in US dollars for that line item."),
        ("value_text", "Text value", "Raw XBRL text of the fact (usually mirrors value_num for amounts)."),
        ("context_ref", "XBRL context reference", "Identifies reporting period and scenario (instant vs duration) in the source filing."),
        ("unit_ref", "Unit reference", "Currency or unit of measure — typically USD."),
    ]
    for abbr, full, defn in columns:
        add("C — Dataset column names", abbr, full, defn)

    # ── D. Reporting forms in this extract ───────────────────────────────────
    for form in sorted(df["reporting_form"].dropna().unique()):
        code = str(form).replace("FFIEC ", "").strip()
        desc = {
            "031": "Consolidated report for banks with foreign offices or complex structures.",
            "032": "Consolidated report variant.",
            "041": "Report for smaller domestic banks — community bank form.",
            "010": "Specialized / non-bank or thrift-related form in panel.",
            "012": "Specialized form variant.",
            "014": "Specialized form variant.",
        }.get(code, "FFIEC Call Report form type filed by this institution.")
        add("D — FFIEC reporting forms", str(form), f"FFIEC Call Report form {code}", desc)

    # ── E. Enumerated field values ────────────────────────────────────────────
    add("E — Field values", "loan_or_lease", "Loan or lease line",
        "mdrm_category value: core loan/lease balance sheet lines from Call Report.")
    add("E — Field values", "schedule_rc_c", "Schedule RC-C line",
        "mdrm_category value: detail lines from Schedule RC-C (loans & leases by category).")
    add("E — Field values", "F", "Financial line item",
        "item_type = F: dollar balance reported on the Call Report.")
    add("E — Field values", "D", "Derived / memorandum",
        "item_type = D: derived, memo, or non-primary line in the taxonomy.")
    add("E — Field values", "TX", "Texas", "Geographic filter: all institutions in this dataset are Texas banks.")
    add("E — Field values", "USD", "US dollars", "unit_ref: all monetary amounts are US dollars.")

    # ── F. Organizations & systems ────────────────────────────────────────────
    orgs = [
        ("FFIEC", "Federal Financial Institutions Examination Council",
         "US interagency body; publishes Call Report data via the Central Data Repository."),
        ("CDR", "Central Data Repository", "FFIEC system at cdr.ffiec.gov — source of XBRL Call Report filings."),
        ("PWS", "Public Web Service", "FFIEC REST API used to download reporting periods, panel, and facsimiles."),
        ("MDRM", "Micro Data Reference Manual",
         "Federal Reserve dictionary mapping RCON/RCFD codes to names and definitions."),
        ("RSSD", "Regulatory Reporting System ID",
         "Federal Reserve permanent numeric ID for a depository institution (stored as id_rssd)."),
        ("XBRL", "eXtensible Business Reporting Language",
         "Machine-readable format of Call Report filings parsed into texas_xbrl_facts.csv."),
        ("Fed", "Federal Reserve Board", "Maintains MDRM dictionary and banking regulatory data standards."),
        ("FDIC", "Federal Deposit Insurance Corporation", "Publishes Schedule RC-C instructions; insures deposits."),
        ("OCC", "Office of the Comptroller of the Currency", "Charters and supervises national banks."),
        ("UBPR", "Uniform Bank Performance Report",
         "Separate peer-ratio report — not included in this Texas extract (different API)."),
        ("ICP", "Ideal Customer Profile",
         "In this workbook: Texas community banks with $500M–$2B total assets."),
    ]
    for abbr, full, defn in orgs:
        add("F — Organizations & systems", abbr, full, defn)

    # ── G. Borrower product abbreviations (EDA taxonomy) ─────────────────────
    for key, (label, parts) in MIX_DEF.items():
        codes = []
        for p in parts:
            codes.extend(PRODUCT_LINES.get(p, []))
        code_str = ", ".join(c for c in codes if c) or "RCON1766 (profiles)"
        add("G — Borrower product codes", key.upper(), label,
            f"Internal EDA bucket key '{key}'. Mapped from MDRM: {code_str}. Used in portfolio mix % calculations.")

    products = [
        ("CRE", "Commercial real estate", "Income-producing or owner-occupied property-secured loans."),
        ("C&I", "Commercial & industrial", "Business operating loans, term debt, and working-capital facilities."),
        ("MF", "Multifamily", "Apartment and 5+ unit residential investment property loans (RCON1460)."),
        ("HELOC", "Home equity line of credit", "Revolving credit secured by 1–4 family residential (RCON1797)."),
        ("RE", "Real estate", "Loans secured by real property (residential or commercial)."),
        ("1–4", "One-to-four family residential", "Single-family and small residential mortgage loans."),
        ("L&D", "Land development", "Construction and land development loans (RCONF159)."),
        ("OO", "Owner-occupied", "CRE occupied by the borrower's own business (RCONF160)."),
        ("NOO", "Non-owner-occupied / investor", "Income-property CRE (RCONF161)."),
        ("Ag", "Agricultural", "Farm production and farmland-secured loans."),
    ]
    for abbr, full, defn in products:
        add("G — Borrower product codes", abbr, full, defn)

    # ── H. EDA metrics & statistical terms ───────────────────────────────────
    metrics = [
        ("HHI", "Herfindahl-Hirschman Index",
         "Market concentration measure. HHI × 10,000 shown in EDA; >1,500 = highly concentrated on DOJ scale."),
        ("Gini", "Gini coefficient",
         "Inequality of balances across banks (0 = equal, 100 = one bank holds all). Shown as Gini × 100 in EDA."),
        ("P25", "25th percentile", "25% of active lenders hold less than this exposure amount or mix %."),
        ("P75", "75th percentile", "75% of active lenders hold less than this exposure amount or mix %."),
        ("QoQ", "Quarter over quarter", "Percentage change from prior reporting period to latest period."),
        ("mix_%", "Portfolio mix percentage", "Product balance ÷ total loans (RCON2122) × 100 for one bank."),
        ("spec.", "Specialist", f"Bank with ≥{int(SPECIALIST_PCT)}% of total loans in one product category."),
        ("≥8%", "Specialist threshold", f"Portfolio share ≥{int(SPECIALIST_PCT)}% — first filter for borrower outreach lists."),
        ("≥15%", "Deep specialist threshold", "Portfolio share ≥15% — bank is heavily oriented to that product."),
        ("Availability %", "Lender availability score",
         "Percentage of Texas banks with any non-zero balance in the product category."),
        ("Top 5 share %", "Top-five concentration",
         "Combined market share of the five largest banks in a product category."),
        ("90+ past-due %", "Serious delinquency ratio",
         "Balances 90+ days past due ÷ total loans — portfolio-level lender health signal."),
        ("$M", "Millions of US dollars", "Monetary unit in EDA tables (values rounded to nearest whole number)."),
        ("$B", "Billions of US dollars", "Monetary unit for statewide aggregate totals."),
        ("Portfolio-style", "CRE + C&I portfolio lender",
         "Bank where investor CRE + owner-occupied CRE + C&I exceed 50% of total loans."),
    ]
    for abbr, full, defn in metrics:
        add("H — EDA metrics & statistics", abbr, full, defn)

    # ── I. Internal field keys (wide-table / profile joins) ───────────────────
    internal = [
        ("total_loans_gross", "Total loans (gross)", "RCON2122 / RCFD2122 — denominator for all mix % calculations."),
        ("mix_inv_usd", "Investor CRE balance", "RCONF161 + RCONF162 balances in USD."),
        ("mix_ci_usd", "C&I balance", "RCON1766 from texas_bank_profiles_latest.csv."),
        ("mix_mf_usd", "Multifamily balance", "RCON1460 balance in USD."),
        ("mix_own_usd", "Owner-occupied CRE balance", "RCONF160 balance in USD."),
        ("mix_con_usd", "Construction balance", "RCONF158 + RCONF159 balances in USD."),
        ("mix_cons_usd", "Consumer balance", "RCON1545 + RCON1583 balances in USD."),
        ("mix_res_usd", "1–4 family residential balance", "RCON1403 balance in USD."),
        ("mix_heloc_usd", "HELOC balance", "RCON1797 balance in USD."),
        ("mix_farm_usd", "Farmland balance", "RCON1420 balance in USD."),
        ("mix_ag_usd", "Ag production balance", "RCON1590 balance in USD."),
        ("mix_lease_usd", "Lease financing balance", "RCON1754 + RCONF163 balances in USD."),
        ("past_due_90_plus", "90+ day past-due balance", "From bank profiles — total seriously delinquent loans."),
        ("icp_fit", "ICP classification", "Flag: ICP ($500M–$2B) vs outside band based on total_assets."),
        ("total_assets", "Total assets", "Bank size from profiles (RCON2170); used for ICP segmentation."),
        ("form_segment", "Form size segment", "Community bank (FFIEC 041) vs regional/larger (031/032)."),
    ]
    for abbr, full, defn in internal:
        add("I — Internal EDA field keys", abbr, full, defn)

    # ── J. Schedule & report references ───────────────────────────────────────
    schedules = [
        ("RC-C", "Schedule RC-C — Loans & leases",
         "Call Report schedule listing loan categories by collateral and purpose; source of most MDRM codes here."),
        ("RC-C Part I", "RC-C loans held in portfolio",
         "Primary loan category breakdown (construction, CRE, consumer, etc.)."),
        ("Call Report", "Consolidated Reports of Condition and Income",
         "Quarterly regulatory financial filing (FFIEC 031/041) filed by US banks."),
        ("Facsimile", "XBRL facsimile", "Official electronic copy of a filed Call Report downloaded via FFIEC API."),
        ("Panel of Reporters", "FFIEC reporter panel",
         "List of institutions expected to file for each reporting period."),
    ]
    for abbr, full, defn in schedules:
        add("J — Regulatory schedules", abbr, full, defn)

    out = pd.DataFrame(rows)
    return out


def write_abbreviations_sheet(ws, glossary: pd.DataFrame) -> None:
    ws.cell(row=1, column=1, value="Abbreviations & glossary — texas_loans_summary").font = BANNER_FONT
    ws.cell(row=1, column=1).fill = BANNER_FILL
    ws.merge_cells("A1:D1")
    ws.cell(
        row=2, column=1,
        value=(
            f"Complete reference for codes, acronyms, and terms | "
            f"{len(glossary)} entries | Built {datetime.now():%Y-%m-%d}"
        ),
    ).font = Font(italic=True, size=10)
    ws.merge_cells("A2:D2")

    headers = list(glossary.columns)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")

    section_fill = PatternFill("solid", fgColor="E7E6E6")
    current_section = None
    row_idx = 5
    for _, rec in glossary.iterrows():
        if rec["Section"] != current_section:
            current_section = rec["Section"]
            c = ws.cell(row=row_idx, column=1, value=current_section)
            c.font = Font(bold=True, size=11)
            c.fill = section_fill
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            row_idx += 1
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=rec[key])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 72
    ws.freeze_panes = "A5"


SECTION_TITLES = {
    "Borrower-perspective EDA — texas_loans_summary",
    "Methodology, assumptions & limitations",
    "ICP specialists by product (≥8%)",
    "ICP bank shortlist — top 30 by investor CRE portfolio share",
}


def _is_section_title(label: str) -> bool:
    if not label:
        return False
    if label in SECTION_TITLES or label.startswith("Top specialists"):
        return True
    if label.startswith("§ "):
        return True
    return bool(label and label[0].isdigit() and "." in label[:4])


def write_table(ws, start_row: int, table: pd.DataFrame) -> int:
    for j, col in enumerate(table.columns, 1):
        c = ws.cell(row=start_row + 1, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True)
    for i, row in enumerate(table.itertuples(index=False), start_row + 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    return start_row + len(table) + 2


def write_narrative_sheet(ws, rows: list, banner: str, subtitle: str, section_style=None) -> None:
    ws.cell(row=1, column=1, value=banner).font = BANNER_FONT
    ws.cell(row=1, column=1).fill = BANNER_FILL
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=10)
    ws.merge_cells("A2:F2")

    sec_font = section_style or SECTION_FONT
    sec_fill = PROV_SECTION if section_style is None else SECTION_FILL

    start_row = 3
    for label, value in rows:
        if label == "__table__" and isinstance(value, pd.DataFrame):
            start_row = write_table(ws, start_row, value)
            start_row += 1
            continue
        if label == "__explanation__":
            start_row += 1
            title_cell = ws.cell(row=start_row, column=1, value="What this table shows")
            title_cell.font = EXPLANATION_LABEL_FONT
            ws.merge_cells(
                start_row=start_row, start_column=1, end_row=start_row, end_column=6,
            )
            start_row += 1
            body_cell = ws.cell(row=start_row, column=1, value=value)
            body_cell.font = Font(size=10)
            body_cell.alignment = Alignment(wrap_text=True, vertical="top")
            body_cell.fill = EXPLANATION_FILL
            ws.merge_cells(
                start_row=start_row, start_column=1, end_row=start_row, end_column=6,
            )
            ws.row_dimensions[start_row].height = max(45, 15 * (1 + len(str(value)) // 120))
            start_row += 1
            continue
        if label == "" and value == "":
            start_row += 1
            continue
        if _is_section_title(label):
            cell = ws.cell(row=start_row + 1, column=1, value=label)
            cell.font = sec_font
            cell.fill = sec_fill
            ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=6)
            start_row += 1
            continue
        ws.cell(row=start_row + 1, column=1, value=label)
        val_cell = ws.cell(row=start_row + 1, column=2, value=value)
        val_cell.alignment = Alignment(wrap_text=True, vertical="top")
        start_row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 92


def write_insights_sheet(ws, insights: list[dict]) -> None:
    ws.cell(row=1, column=1, value="Borrower Insights — texas_loans_summary").font = BANNER_FONT
    ws.cell(row=1, column=1).fill = BANNER_FILL
    ws.merge_cells("A1:H1")
    ws.cell(row=2, column=1, value=(
        f"Documented findings with plain-English explanations | Built {datetime.now():%Y-%m-%d} | "
        f"{len(insights)} insights"
    )).font = Font(italic=True, size=10)
    ws.merge_cells("A2:H2")

    headers = [
        "ID", "Priority", "Theme", "EDA section", "Finding",
        "What this means for borrowers", "Evidence", "Recommended action",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True)

    for i, ins in enumerate(insights, 5):
        row_vals = [
            ins["id"], ins["priority"], ins["theme"], ins.get("section", ""),
            ins["insight"], ins.get("explanation", ""), ins["evidence"], ins["borrower_action"],
        ]
        fill = INSIGHT_HIGH if ins["priority"] == "High" else INSIGHT_MED if ins["priority"] == "Medium" else None
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill

    widths = [8, 10, 16, 12, 42, 42, 28, 32]
    for col, width in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = width


def main() -> int:
    if not INPUT.exists():
        print(f"Missing {INPUT}")
        print("Run: python ONLY_TEXAS_SINCE_2025/extract_texas_loans.py --summary")
        return 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_COPY.parent.mkdir(parents=True, exist_ok=True)
    tmp = OUTPUT.with_suffix(".build.xlsx")

    print(f"Loading {INPUT} …")
    df = pd.read_csv(INPUT)
    print(f"  {len(df):,} rows × {len(df.columns)} columns")

    print("Running exhaustive borrower EDA …")
    eda_rows, insights = run_analysis(df)

    print("Building data provenance documentation …")
    prov_rows = build_provenance_rows(df)

    print("Building abbreviations glossary …")
    abbrev_glossary = build_abbreviations_glossary(df)

    print(f"Writing workbook → {OUTPUT} …")
    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="texas_loans_summary", index=False)
        pd.DataFrame({"": []}).to_excel(writer, sheet_name="EDA", index=False)
        pd.DataFrame({"": []}).to_excel(writer, sheet_name="Insights", index=False)
        pd.DataFrame({"": []}).to_excel(writer, sheet_name="Data Provenance", index=False)
        pd.DataFrame({"": []}).to_excel(writer, sheet_name="Abbreviations", index=False)

    from openpyxl import load_workbook
    import shutil

    wb = load_workbook(tmp)
    write_narrative_sheet(
        wb["EDA"], eda_rows,
        "EDA — Exhaustive borrower-perspective analysis",
        f"{len(insights)} documented insights on Insights sheet | Latest quarter emphasized",
    )
    write_insights_sheet(wb["Insights"], insights)
    write_narrative_sheet(
        wb["Data Provenance"], prov_rows,
        "Data Provenance — How texas_loans_summary was retrieved and built",
        "FFIEC Public Web Service API → XBRL archive → MDRM-labeled CSV → this workbook",
        section_style=PROV_SECTION_FONT,
    )
    write_abbreviations_sheet(wb["Abbreviations"], abbrev_glossary)

    for sheet in ("texas_loans_summary",):
        for cell in wb[sheet][1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    wb.save(tmp)
    tmp.replace(OUTPUT)
    shutil.copy2(OUTPUT, OUTPUT_COPY)

    print(f"Done.")
    print(f"  {OUTPUT}")
    print(f"  {OUTPUT_COPY}")
    print(f"  Sheets: texas_loans_summary ({len(df):,} rows), EDA, Insights ({len(insights)}), "
          f"Data Provenance, Abbreviations ({len(abbrev_glossary)} entries)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
