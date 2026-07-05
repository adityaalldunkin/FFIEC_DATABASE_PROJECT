#!/usr/bin/env python3
"""Build All files exploratory data analysis.xlsx with borrower-focused EDA."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from eda_insights import ROOT, build_eda
from eda_extras import build_provenance_rows

OUTPUT = ROOT / "All files exploratory data analysis.xlsx"

SOURCE_FILES: list[Path] = [
    ROOT / "institutions_definitions.csv",
    ROOT / "events_definitions.csv",
    ROOT / "institutions.csv",
    ROOT / "locations.csv",
    ROOT / "locations_definitions.csv",
    ROOT / "sod_variables_definitions.csv",
    ROOT / "texas_filings.csv",
    ROOT / "texas_institutions.csv",
    ROOT / "texas_loans_labeled.csv",
    ROOT / "texas_loan_products_mdrm_catalog.csv",
    ROOT / "texas_mdrm_loan_taxonomy.csv",
    ROOT / "26.05.18.All.Loan.Types.UBPR.Reference.xlsx",
    ROOT / "All Financial Reports.xlsx",
    ROOT / "texas_mdrm_loan_taxonomy.xlsx",
    ROOT / "ONLY_TEXAS_SINCE_2025/analysis/asset_band_counts.csv",
    ROOT / "ONLY_TEXAS_SINCE_2025/analysis/lenni_icp_prospect_list.csv",
    ROOT / "ONLY_TEXAS_SINCE_2025/analysis/lenni_market_segments.csv",
    ROOT / "ONLY_TEXAS_SINCE_2025/analysis/summary_statistics_latest.csv",
]

MAX_DATA_ROWS = 5000
EXCEL_SHEET_MAX = 31

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=12)
BANNER_FILL = PatternFill("solid", fgColor="2E75B6")
BANNER_FONT = Font(bold=True, color="FFFFFF", size=11)

SECTION_TITLES = {
    "Dataset overview",
    "Schema",
    "Content profile",
    "Content themes",
    "Key fields for borrowers",
    "Key insights",
    "Column profile",
    "Numeric summary",
    "Data quality and caveats",
    "Borrower decision guide",
    "How this connects to other files",
    "Detailed insights catalog — everything you can learn from this file",
    "Abbreviations and terms used in this file",
    "Workbook sheets",
    "Texas coverage by borrower product category",
    "Investor CRE — sub-types with widest Texas bank coverage",
    "Borrower View — product guide (full table)",
    "Loan products by UBPR category",
    "Template library (build sequence)",
    "Texas banks by asset band",
    "Market segments",
    "Portfolio summary (ICP banks)",
    "Top 10 ICP banks by CRE concentration (% of total loans)",
    "Top 10 ICP banks by C&I concentration (% of total loans)",
    "Distribution of key metrics (Texas banks with data)",
    "Texas active banks by asset band",
    "Texas active banks by charter class (BKCLASS)",
    "Top Texas cities by active bank count",
    "Top 15 Texas cities by branch count",
    "All location field definitions",
    "Field definitions (sample)",
    "Sample event fields (merger / acquisition related)",
    "Filings per reporting period",
    "Filing status",
    "Top cities by bank count",
    "Rows by regulatory category",
    "Rows by Call Report form (bank size)",
    "Rows by reporting period",
    "All catalog codes by category",
    "Texas-observed codes by category",
    "Executive summary — three data pipelines",
    "Source systems and APIs",
    "Phase A — Texas Call Report download (FFIEC API)",
    "Phase B — MDRM labeling (loan product names)",
    "Phase C — Lenni taxonomy & borrower mapping",
    "Phase D — FDIC institution & branch enrichment",
    "Phase E — Lenni EDA analysis outputs",
    "Phase F — Reference workbooks (manual / internal)",
    "File lineage — each workbook file traced to source",
    "Credentials, environment, and rerun commands",
    "Important limitations (all pipelines)",
    "Abbreviations used in this provenance document",
}


def _tab_base(filename: str, suffix_len: int = 4) -> str:
    """Filename portion of tab so base + suffix fits in 31 Excel characters."""
    max_base = EXCEL_SHEET_MAX - suffix_len
    if len(filename) <= max_base:
        return filename
    ext = Path(filename).suffix
    stem = Path(filename).stem
    room = max_base - len(ext)
    return f"{stem[:room]}{ext}"


def sheet_name_for_file(filename: str, kind: str) -> str:
    """Tab: `{filename}_dat` or `{filename}_EDA` (full name when ≤27 chars).

    Row 1 banner on each sheet always shows the complete filename and path.
    """
    kind_suffix = "_dat" if kind == "data" else "_EDA"
    return f"{_tab_base(filename, len(kind_suffix))}{kind_suffix}"


def human_size(num_bytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if num_bytes < 1024 or unit == "GB":
            return f"{num_bytes:.1f} {unit}" if unit != "B" else f"{num_bytes} B"
        num_bytes /= 1024
    return f"{num_bytes:.1f} GB"


def csv_skiprows(path: Path) -> int:
    if path.suffix.lower() != ".csv":
        return 0
    first = path.open(encoding="utf-8", errors="replace").readline().strip()
    title_rows = {
        "institutions_definitions",
        "events_definitions",
        "locations_definitions",
        "sod_variables_definitions",
    }
    return 1 if first in title_rows else 0


def read_source(path: Path) -> tuple[pd.DataFrame, dict]:
    meta: dict = {
        "path": str(path.relative_to(ROOT)),
        "filename": path.name,
        "format": path.suffix.lower(),
    }
    if not path.exists():
        meta["error"] = "File not found"
        return pd.DataFrame(), meta

    meta["size_bytes"] = path.stat().st_size
    meta["size_human"] = human_size(meta["size_bytes"])
    sk = csv_skiprows(path)

    if path.suffix.lower() == ".csv":
        nrows = None
        if path.stat().st_size > 5_000_000:
            with path.open(encoding="utf-8", errors="replace") as fh:
                approx = sum(1 for _ in fh) - 1 - sk
            meta["approx_rows"] = approx
            nrows = MAX_DATA_ROWS
            meta["data_note"] = (
                f"Full file has ~{approx:,} rows. Data sheet shows first {MAX_DATA_ROWS:,} rows. "
                f"EDA uses full-file statistics where noted."
            )
        df = pd.read_csv(path, skiprows=sk or None, nrows=nrows, low_memory=False)
        meta["skiprows"] = sk
    elif path.suffix.lower() == ".xlsx":
        xl = pd.ExcelFile(path)
        meta["xlsx_sheets"] = xl.sheet_names
        frames: list[pd.DataFrame] = []
        for sheet in xl.sheet_names:
            part = pd.read_excel(path, sheet_name=sheet)
            if part.empty:
                continue
            part = part.copy()
            part.insert(0, "__source_sheet__", sheet)
            frames.append(part)
        if not frames:
            meta["error"] = "All Excel sheets are empty"
            df = pd.DataFrame()
        else:
            df = pd.concat(frames, ignore_index=True, sort=False)
            meta["xlsx_sheets_with_data"] = [str(f["__source_sheet__"].iloc[0]) for f in frames]
            meta["data_note"] = (
                f"Multi-sheet workbook. Data sheet combines {len(frames)} non-empty sheets "
                f"({', '.join(meta['xlsx_sheets_with_data'])}), prefixed with __source_sheet__."
            )
            if len(df) > MAX_DATA_ROWS:
                meta["data_note"] += f" Showing first {MAX_DATA_ROWS:,} combined rows."
                df = df.head(MAX_DATA_ROWS)
    else:
        meta["error"] = f"Unsupported format: {path.suffix}"
        return pd.DataFrame(), meta

    meta["rows_in_sheet"] = len(df)
    meta["cols"] = len(df.columns)
    meta["columns"] = list(df.columns)
    return df, meta


def add_banner(ws, full_filename: str, rel_path: str, kind: str) -> None:
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    c1 = ws.cell(row=1, column=1, value=f"Source file: {full_filename}")
    c1.font = BANNER_FONT
    c1.fill = BANNER_FILL
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c2 = ws.cell(row=2, column=1, value=f"Path: {rel_path}  |  Sheet type: {kind}  |  Built: {datetime.now():%Y-%m-%d}")
    c2.font = Font(italic=True, size=10)
    c2.alignment = Alignment(horizontal="left")


def write_eda_sheet(writer: pd.ExcelWriter, sheet_name: str, eda_rows: list, filename: str, rel_path: str) -> None:
    pd.DataFrame({"": [""]}).to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
    ws = writer.sheets[sheet_name]
    add_banner(ws, filename, rel_path, "EDA insights")

    start_row = 2
    for label, value in eda_rows:
        if label == "__table__" and isinstance(value, pd.DataFrame):
            table = value.copy()
            table.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += len(table) + 3
            continue
        if label == "" and value == "":
            start_row += 1
            continue
        if label in SECTION_TITLES or label.startswith("Value counts"):
            cell = ws.cell(row=start_row + 1, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=2)
            start_row += 1
            continue
        ws.cell(row=start_row + 1, column=1, value=label)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            ws.cell(row=start_row + 1, column=2, value=value)
        else:
            c = ws.cell(row=start_row + 1, column=2, value=str(value))
            c.alignment = Alignment(wrap_text=True, vertical="top")
        start_row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90


def style_header_row(ws, row: int) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True)


def autofit_columns(ws, max_width: int = 55, start_col: int = 1) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        if col_idx < start_col:
            continue
        length = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), max_width)


def build_workbook() -> None:
    index_rows = []
    used_names: set[str] = set()

    def unique_sheet(filename: str, kind: str) -> str:
        base = sheet_name_for_file(filename, kind)
        if base not in used_names:
            used_names.add(base)
            return base
        # disambiguate with path hash fragment
        n = 2
        while True:
            alt = sheet_name_for_file(f"{filename} ({n})", kind)
            if len(alt) > EXCEL_SHEET_MAX:
                alt = alt[:EXCEL_SHEET_MAX]
            if alt not in used_names:
                used_names.add(alt)
                return alt
            n += 1

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for i, path in enumerate(SOURCE_FILES, start=1):
            filename = path.name
            rel_path = str(path.relative_to(ROOT))
            data_sheet = unique_sheet(filename, "data")
            eda_sheet = unique_sheet(filename, "EDA")

            df, meta = read_source(path)

            if not df.empty:
                df.to_excel(writer, sheet_name=data_sheet, index=False, startrow=2)
                ws = writer.sheets[data_sheet]
                add_banner(ws, filename, rel_path, "original data")
                style_header_row(ws, row=3)
                autofit_columns(ws, start_col=1)
            else:
                note = pd.DataFrame({
                    "message": [
                        meta.get("error", "No tabular data loaded."),
                        f"Full source path: {rel_path}",
                        f"Workbook sheets: {', '.join(meta.get('xlsx_sheets', []))}",
                    ]
                })
                note.to_excel(writer, sheet_name=data_sheet, index=False, startrow=2)
                add_banner(writer.sheets[data_sheet], filename, rel_path, "original data")
                style_header_row(writer.sheets[data_sheet], row=3)

            eda_rows = build_eda(path, df, meta)
            write_eda_sheet(writer, eda_sheet, eda_rows, filename, rel_path)

            index_rows.append({
                "#": i,
                "Source file (full name)": filename,
                "Source path": rel_path,
                "Data sheet tab": data_sheet,
                "EDA sheet tab": eda_sheet,
                "Sheet tab note": (
                    "Full filename in tab when ≤31 chars; otherwise truncated — "
                    "see row 1 banner on each sheet for complete name."
                ),
                "EDA status": "Complete",
                "Rows in data sheet": meta.get("rows_in_sheet", 0),
                "Columns": meta.get("cols", 0),
                "File size": meta.get("size_human", ""),
                "Notes": meta.get("data_note", meta.get("error", "")),
            })

        guide = pd.DataFrame([
            {
                "Topic": "Who this workbook is for",
                "Detail": "Texas commercial real estate and business borrowers using Lenni to find community banks by loan type, geography, and portfolio fit.",
            },
            {
                "Topic": "How sheets are named",
                "Detail": "Each file has two tabs: '{filename}_dat' and '{filename}_EDA'. Row 1 banner shows complete filename and path.",
            },
            {
                "Topic": "New: Detailed insights catalog",
                "Detail": "Every EDA sheet ends with a table of ALL insights you can generate, plus an abbreviations glossary.",
            },
            {
                "Topic": "New: Data provenance",
                "Detail": "See sheet 'How All Data Was Pulled' for step-by-step documentation of every API, script, and pipeline phase.",
            },
            {
                "Topic": "Recommended reading order",
                "Detail": "1) How All Data Was Pulled → 2) texas_mdrm_loan_taxonomy.xlsx EDA → 3) lenni_icp_prospect_list.csv EDA → 4) texas_loans_labeled.csv EDA",
            },
            {
                "Topic": "Rebuild command",
                "Detail": "python build_eda_workbook.py",
            },
        ])
        guide.to_excel(writer, sheet_name="README", index=False)
        style_header_row(writer.sheets["README"], row=1)

        # Data provenance sheet
        prov_rows = build_provenance_rows()
        prov_sheet = "How All Data Was Pulled"
        pd.DataFrame({"": [""]}).to_excel(writer, sheet_name=prov_sheet, index=False, startrow=1)
        write_eda_sheet(
            writer,
            prov_sheet,
            prov_rows,
            "All source files (combined provenance)",
            "See phases below",
        )
        # Override banner for provenance
        ws_prov = writer.sheets[prov_sheet]
        ws_prov.cell(row=1, column=1, value="How all original data was pulled — complete provenance guide")
        ws_prov.cell(row=2, column=1, value="Covers FFIEC API, Federal Reserve MDRM, FDIC BankFind, and Lenni analysis scripts")

        index_df = pd.DataFrame(index_rows)
        index_df.to_excel(writer, sheet_name="Index", index=False)
        style_header_row(writer.sheets["Index"], row=1)
        autofit_columns(writer.sheets["Index"], max_width=45)

        wb = writer.book
        wb.move_sheet("README", offset=-len(wb.sheetnames) + 1)
        wb.move_sheet("Index", offset=-len(wb.sheetnames) + 1)
        wb.move_sheet(prov_sheet, offset=-len(wb.sheetnames) + 1)

    print(f"Wrote {OUTPUT} ({human_size(OUTPUT.stat().st_size)})")
    names = load_workbook(OUTPUT, read_only=True).sheetnames
    print(f"Sheets: {len(names)}")
    for n in names[:6]:
        print(f"  {n}")
    print("  ...")


if __name__ == "__main__":
    build_workbook()
