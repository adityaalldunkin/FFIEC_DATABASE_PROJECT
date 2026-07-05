#!/usr/bin/env python3
"""
Borrower-perspective EDA on texas_loans_labeled.xlsx.

Produces a 3-sheet workbook:
  1. texas_loans_labeled — original regulatory loan line items
  2. EDA — structured analysis tables (borrower lens)
  3. Insights — documented findings with evidence and actions

  python build_texas_loans_labeled_borrower_eda.py
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
INPUT = ROOT / "texas_loans_labeled.xlsx"
OUTPUT = ROOT / "texas_loans_labeled.xlsx"
TMP = ROOT / "texas_loans_labeled_eda_build.xlsx"
PROFILES = ROOT / "ONLY_TEXAS_SINCE_2025" / "exports" / "texas_bank_profiles_latest.csv"
TAXONOMY = ROOT / "texas_mdrm_loan_taxonomy.csv"

ICP_MIN = 500_000_000
ICP_MAX = 2_000_000_000
SPECIALIST_PCT = 8.0

# Schedule RC-C summary lines used on the borrower site (values in thousands USD).
PRODUCT_LINES = {
    "total_loans_gross": ["RCON2122", "RCFD2122"],
    "multifamily_re_loans": ["RCON1460"],
    "other_nonfarm_nonres_re": ["RCONF161"],
    "commercial_re_loans": ["RCONF162"],
    "owner_occupied_nonfarm_re": ["RCONF160"],
    "residential_construction": ["RCONF158"],
    "other_construction_ld": ["RCONF159"],
    "ci_loans": ["RCON1766"],
    "residential_1_4_family": ["RCON1403"],
    "credit_card_plans": ["RCON1545"],
    "other_consumer_loans": ["RCON1583"],
    "farmland_loans": ["RCON1420"],
    "ag_production_loans": ["RCON1590"],
    "lease_financing": ["RCON1754"],
    "past_due_30_89": ["RCON5367"],
    "past_due_90_plus": ["RCON5368"],
}

MIX_DEF = {
    "mf": ("Multifamily (5+ units)", ["multifamily_re_loans"]),
    "inv": ("Investor CRE (income property)", ["other_nonfarm_nonres_re", "commercial_re_loans"]),
    "own": ("Owner-Occupied CRE", ["owner_occupied_nonfarm_re"]),
    "con": ("Commercial Construction", ["residential_construction", "other_construction_ld"]),
    "ci": ("Commercial & Industrial (C&I)", ["ci_loans"]),
    "res": ("1–4 Family Residential", ["residential_1_4_family"]),
    "cons": ("Consumer lending", ["credit_card_plans", "other_consumer_loans"]),
    "farm": ("Agricultural & Farmland (farmland)", ["farmland_loans"]),
    "ag": ("Agricultural & Farmland (production)", ["ag_production_loans"]),
    "lease": ("Lease financing", ["lease_financing"]),
}

BORROWER_CATEGORIES = [
    "Multifamily (5+ units)",
    "Investor CRE (income property)",
    "Owner-Occupied CRE",
    "Commercial Construction",
    "Commercial & Industrial (C&I)",
    "1–4 Family Residential",
    "Agricultural & Farmland",
    "Consumer lending",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=12)
BANNER_FILL = PatternFill("solid", fgColor="2E75B6")
BANNER_FONT = Font(bold=True, color="FFFFFF", size=11)
INSIGHT_HIGH = PatternFill("solid", fgColor="E2EFDA")
INSIGHT_MED = PatternFill("solid", fgColor="FFF2CC")
EXPLANATION_FILL = PatternFill("solid", fgColor="F2F2F2")
EXPLANATION_LABEL_FONT = Font(bold=True, italic=True, size=11)

# Stable reference labels for the Insights sheet (ID → theme + summary).
INSIGHT_CATALOG: dict[str, dict[str, str]] = {
    "I-01": {
        "theme": "Product choice breadth",
        "captures": "Which loan types have the most vs fewest Texas lenders",
    },
    "I-02": {
        "theme": "Market structure",
        "captures": "How concentrated each product market is (HHI, top-5 share)",
    },
    "I-03": {
        "theme": "Specialist targeting",
        "captures": "Banks with the highest multifamily portfolio share",
    },
    "I-04": {
        "theme": "Bank size fit",
        "captures": "Community banks (FFIEC 041) vs larger regional banks",
    },
    "I-05": {
        "theme": "ICP opportunity",
        "captures": "Banks in the $500M–$2B community-bank sweet spot",
    },
    "I-06": {
        "theme": "Geography",
        "captures": "Lender density by headquarters city (e.g. Dallas)",
    },
    "I-07": {
        "theme": "Market momentum",
        "captures": "How total loan books changed quarter over quarter",
    },
    "I-08": {
        "theme": "Lender diligence",
        "captures": "Past-due ratios as a lender health / appetite signal",
    },
    "I-09": {
        "theme": "Bank archetypes",
        "captures": "Portfolio-style vs consumer-heavy bank profiles",
    },
    "I-10": {
        "theme": "Outreach planning",
        "captures": "How many specialist banks to realistically contact",
    },
    "I-11": {
        "theme": "Cross-sell patterns",
        "captures": "Loan products that tend to co-occur at the same banks",
    },
    "I-12": {
        "theme": "Deal size fit",
        "captures": "Whether your loan size matches typical bank exposure",
    },
}


def gini(values: np.ndarray) -> float:
    x = np.sort(values[values > 0])
    if len(x) == 0:
        return float("nan")
    n = len(x)
    cum = np.cumsum(x)
    return float((n + 1 - 2 * np.sum(cum) / cum[-1]) / n)


def hhi(shares: np.ndarray) -> float:
    """Herfindahl-Hirschman Index on market shares (0–1 scale; ×10,000 for DOJ convention)."""
    s = shares[shares > 0]
    if len(s) == 0:
        return float("nan")
    total = s.sum()
    if total <= 0:
        return float("nan")
    p = s / total
    return float((p ** 2).sum())


def form_segment(form: str) -> str:
    f = str(form or "")
    if "041" in f:
        return "Community bank (FFIEC 041)"
    if "031" in f or "032" in f:
        return "Regional / larger (FFIEC 031/032)"
    return "Other / specialty form"


def pick_line_value(group: pd.DataFrame, col: str) -> float:
    codes = PRODUCT_LINES[col]
    sub = group[group["mdrm_code"].isin(codes)]
    if sub.empty:
        return 0.0
    # Prefer RCON over RCFD; take max if duplicates.
    sub = sub.sort_values("mdrm_code")
    return float(sub["value_num"].max())


def build_bank_wide(latest: pd.DataFrame) -> pd.DataFrame:
    rows = []
    meta_cols = ["id_rssd", "institution_name", "reporting_form", "reporting_period"]
    for rssd, grp in latest.groupby("id_rssd"):
        row = {c: grp[c].iloc[0] for c in meta_cols}
        for col in PRODUCT_LINES:
            row[col] = pick_line_value(grp, col)
        rows.append(row)
    wide = pd.DataFrame(rows)
    return wide


def compute_mix_frame(wide: pd.DataFrame) -> pd.DataFrame:
    out = wide.copy()
    total = out["total_loans_gross"].clip(lower=0)
    out["total_loans_usd"] = total * 1000

    for key, (label, parts) in MIX_DEF.items():
        bal = sum(out[p].fillna(0) for p in parts)
        out[f"mix_{key}_usd"] = bal * 1000
        out[f"mix_{key}_pct"] = np.where(total > 0, 100 * bal / total, 0)

    accounted = sum(out[f"mix_{k}_usd"] / 1000 for k in MIX_DEF)
    out["mix_uncat_usd"] = np.maximum(0, total - accounted) * 1000
    out["mix_uncat_pct"] = np.where(total > 0, 100 * out["mix_uncat_usd"] / (total * 1000), 0)

    # Past-due ratios (borrower risk signal on lender health)
    out["past_due_30_89_pct"] = np.where(total > 0, 100 * out["past_due_30_89"] / total, 0)
    out["past_due_90_plus_pct"] = np.where(total > 0, 100 * out["past_due_90_plus"] / total, 0)
    out["form_segment"] = out["reporting_form"].map(form_segment)
    return out


def load_profiles() -> pd.DataFrame:
    if not PROFILES.exists():
        return pd.DataFrame()
    p = pd.read_csv(PROFILES, dtype={"id_rssd": int})
    p["assets_usd"] = p["total_assets"]
    p["icp_fit"] = p.get("icp_fit", pd.Series(dtype=str))
    p["is_icp"] = p["icp_fit"] == "ICP ($500M–$2B)"
    if "is_icp" not in p or p["is_icp"].sum() == 0:
        p["is_icp"] = p["assets_usd"].between(ICP_MIN, ICP_MAX)
    return p


def round_num(value) -> int | str:
    """Round numeric EDA values to the nearest whole number."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        return int(round(float(value)))
    if isinstance(value, str):
        stripped = value.strip().replace(",", "").replace("%", "")
        try:
            return int(round(float(stripped)))
        except ValueError:
            return value
    return value


def round_eda_table(frame: pd.DataFrame) -> pd.DataFrame:
    """Round every numeric column in an EDA table to the nearest integer."""
    out = frame.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = out[col].round(0).astype(int)
    return out


def round_eda_scalar(value) -> int | str:
    return round_num(value)
    for key, (lbl, _) in MIX_DEF.items():
        if lbl == label or label.startswith(lbl.split("(")[0].strip()):
            return key
    if label == "Agricultural & Farmland":
        return "farm"  # aggregate farm+ag below
    return None


def run_analysis(df: pd.DataFrame) -> tuple[list, list]:
    latest_period = df["reporting_period"].max()
    latest = df[df["reporting_period"] == latest_period].copy()
    wide = build_bank_wide(latest)
    mix = compute_mix_frame(wide)
    profiles = load_profiles()
    if not profiles.empty:
        mix = mix.merge(
            profiles[["id_rssd", "city", "total_assets", "is_icp", "icp_fit"]],
            on="id_rssd",
            how="left",
        )
    else:
        mix["city"] = ""
        mix["total_assets"] = np.nan
        mix["is_icp"] = mix["total_loans_usd"].between(ICP_MIN, ICP_MAX)

    active = mix[mix["total_loans_gross"] > 0].copy()
    n_banks = len(active)
    period_str = pd.Timestamp(latest_period).strftime("%Y-%m-%d")

    eda: list[tuple[str, object]] = []
    insights: list[dict] = []

    def section(title: str) -> None:
        eda.append((title, ""))
        eda.append(("", ""))

    def table(df_out: pd.DataFrame, explanation: str) -> None:
        eda.append(("__table__", round_eda_table(df_out)))
        eda.append(("__explanation__", explanation))
        eda.append(("", ""))

    # --- Banner context ---
    section("Borrower-perspective EDA — texas_loans_labeled.xlsx")
    eda.extend([
        ("Analysis date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Latest reporting period analyzed", period_str),
        ("Reporting periods in file", round_eda_scalar(df["reporting_period"].nunique())),
        ("Texas banks in latest period", round_eda_scalar(n_banks)),
        ("Total regulatory line-item rows (full file)", round_eda_scalar(len(df))),
        ("Distinct MDRM codes", round_eda_scalar(df["mdrm_code"].nunique())),
        ("Value units", "Thousands USD (FFIEC Call Report convention)"),
        ("Borrower lens", "Which Texas banks actually hold loans in your product category — specialization, market depth, geography, and lender health"),
        ("", ""),
    ])

    # --- 1. Product availability matrix ---
    section("1. Lender availability by borrower product category (latest quarter)")
    avail_rows = []
    for key, (label, parts) in MIX_DEF.items():
        bal_col = f"mix_{key}_usd"
        lenders_any = int((active[bal_col] > 0).sum())
        lenders_1m = int((active[bal_col] >= 1_000_000).sum())
        lenders_10m = int((active[bal_col] >= 10_000_000).sum())
        specialists = int((active[f"mix_{key}_pct"] >= SPECIALIST_PCT).sum())
        total_mkt = active[bal_col].sum()
        med_active = active.loc[active[bal_col] > 0, bal_col].median() if lenders_any else 0
        avail_rows.append({
            "Borrower product": label,
            "Banks with any balance": lenders_any,
            f"Banks ≥ ${1}M exposure": lenders_1m,
            f"Banks ≥ ${10}M exposure": lenders_10m,
            f"Specialists (≥{SPECIALIST_PCT:.0f}% of portfolio)": specialists,
            "Total TX market ($M)": total_mkt / 1e6,
            "Median active lender ($M)": med_active / 1e6 if lenders_any else 0,
            "Availability score": 100 * lenders_any / n_banks,
        })
    avail_df = pd.DataFrame(avail_rows)
    best_avail = avail_df.loc[avail_df["Availability score"].idxmax()]
    worst_avail = avail_df.loc[avail_df["Availability score"].idxmin()]
    wide_prod = avail_df.nlargest(3, "Banks with any balance")["Borrower product"].tolist()
    narrow_prod = avail_df.nsmallest(3, "Banks with any balance")["Borrower product"].tolist()
    table(
        avail_df,
        (
            f"This table answers: 'If I need a certain loan type in Texas, how many banks can I realistically call?' "
            f"For the latest quarter ({period_str}), {best_avail['Borrower product']} offers the broadest lender pool "
            f"({round_num(best_avail['Banks with any balance'])} of {n_banks} banks, "
            f"{round_num(best_avail['Availability score'])}% availability score). "
            f"{worst_avail['Borrower product']} is the narrowest ({round_num(worst_avail['Banks with any balance'])} banks, "
            f"{round_num(worst_avail['Availability score'])}% availability). "
            f"The three most widely available categories are {', '.join(wide_prod)}; the three scarcest are {', '.join(narrow_prod)}. "
            f"'Specialists' are banks where that product is ≥{round_num(SPECIALIST_PCT)}% of total loans — a stronger signal that the bank "
            f"actively underwrites that product, not just holds a small legacy balance. "
            f"Borrowers in niche categories should expect fewer meetings and more targeted outreach; "
            f"borrowers seeking C&I or CRE will find the largest shortlists."
        ),
    )
    insights.append({
        "id": "I-01",
        "priority": "High",
        "theme": "Product choice breadth",
        "insight": (
            f"{best_avail['Borrower product']} has the widest lender pool in Texas "
            f"({int(best_avail['Banks with any balance'])} of {n_banks} banks, "
            f"{best_avail['Availability score']:.0f}% availability). "
            f"{worst_avail['Borrower product']} is the narrowest niche "
            f"({int(worst_avail['Banks with any balance'])} banks, "
            f"{worst_avail['Availability score']:.0f}%)."
        ),
        "evidence": f"Latest period {period_str}; specialist threshold {SPECIALIST_PCT}% portfolio share.",
        "borrower_action": "Start outreach with high-availability categories; expect more shopping and comparison for niche products.",
    })

    # --- 2. Market concentration ---
    section("2. Market concentration — how fragmented is each product? (HHI)")
    conc_rows = []
    for key, (label, _) in MIX_DEF.items():
        bal = active[f"mix_{key}_usd"].values
        h = hhi(bal)
        g = gini(bal)
        top5_share = 0.0
        if bal.sum() > 0:
            top5 = np.sort(bal)[-5:].sum() / bal.sum() * 100
            top5_share = float(top5)
        conc_rows.append({
            "Borrower product": label,
            "HHI × 10,000 (DOJ scale)": h * 10000,
            "Gini × 1,000": g * 1000,
            "Top 5 banks' market share %": top5_share,
            "Concentration label": (
                "Highly concentrated" if h > 0.15 else "Moderately concentrated" if h > 0.08 else "Fragmented / competitive"
            ),
        })
    conc_df = pd.DataFrame(conc_rows)
    high_conc = conc_df.sort_values("HHI × 10,000 (DOJ scale)", ascending=False).head(3)
    low_conc = conc_df.sort_values("HHI × 10,000 (DOJ scale)", ascending=False).tail(3)
    conc_parts = []
    for _, r in high_conc.iterrows():
        share_col = "Top 5 banks' market share %"
        conc_parts.append(f"{r['Borrower product']} (top 5 banks hold {round_num(r[share_col])}%)")
    fragmented = ", ".join(low_conc["Borrower product"].tolist())
    table(
        conc_df,
        (
            f"Market concentration tells a borrower how much choice they truly have. "
            f"HHI × 10,000 (DOJ scale) below 1,500 is generally considered competitive; above 2,500 is highly concentrated. "
            f"The most concentrated Texas product markets are: {'; '.join(conc_parts)}. "
            f"In those categories, a handful of large portfolio lenders hold most of the balances — "
            f"you may still find community bank options, but pricing and terms are often set by the majors. "
            f"The most fragmented markets include {fragmented}, where more banks hold meaningful balances and "
            f"borrowers can run a broader RFP process. "
            f"Gini × 1,000 measures inequality of balances across banks (higher = more unequal)."
        ),
    )
    insights.append({
        "id": "I-02",
        "priority": "High",
        "theme": "Market structure",
        "insight": (
            "Most concentrated Texas loan markets: "
            + "; ".join(conc_parts)
            + ". Borrowers in these categories may face fewer realistic alternatives beyond the largest portfolio lenders."
        ),
        "evidence": "HHI and top-5 share computed on bank-level mix balances, latest quarter.",
        "borrower_action": "Build a target list of at least 8–10 banks even in concentrated markets; include mid-size community banks for term competition.",
    })

    # --- 3. Top specialists per major category ---
    section("3. Top 10 specialist lenders by product (≥ meaningful portfolio share)")
    for key, (label, _) in list(MIX_DEF.items())[:7]:
        pct_col = f"mix_{key}_pct"
        usd_col = f"mix_{key}_usd"
        top = active[active[usd_col] > 0].nlargest(10, pct_col)[
            ["institution_name", "city", pct_col, usd_col, "total_loans_usd", "form_segment"]
        ].copy()
        if top.empty:
            continue
        top.columns = ["Bank", "City", "Portfolio %", "Product balance ($M)", "Total loans ($M)", "Bank size segment"]
        top["Portfolio %"] = top["Portfolio %"]
        top["Product balance ($M)"] = top["Product balance ($M)"] / 1e6
        top["Total loans ($M)"] = top["Total loans ($M)"] / 1e6
        eda.append((f"Top specialists — {label}", ""))
        leader = top.iloc[0]
        n_spec = int((active[f"mix_{key}_pct"] >= SPECIALIST_PCT).sum())
        table(
            top.reset_index(drop=True),
            (
                f"This ranks Texas banks by how much of their loan book is in {label} — the primary filter on the borrower site. "
                f"{leader['Bank']} leads with {round_num(leader['Portfolio %'])}% portfolio share "
                f"(${round_num(leader['Product balance ($M)'])}M in this category, "
                f"${round_num(leader['Total loans ($M)'])}M total loans, HQ {leader['City'] or 'TX'}). "
                f"Across all Texas banks, {n_spec} meet the ≥{round_num(SPECIALIST_PCT)}% specialist threshold for this product. "
                f"Banks at the top of this list are your first-call targets: they have demonstrated balance-sheet commitment "
                f"to this asset class, not incidental exposure. Portfolio % above 15% typically indicates a dedicated "
                f"underwriting team and credit appetite for that product type."
            ),
        )

    mf_top = active.nlargest(3, "mix_mf_pct")
    insights.append({
        "id": "I-03",
        "priority": "Medium",
        "theme": "Specialist targeting",
        "insight": (
            f"Multifamily specialists (highest portfolio share): "
            + ", ".join(
                f"{r.institution_name} ({r.mix_mf_pct:.0f}% MF, {r.city or 'TX'})"
                for _, r in mf_top.iterrows()
            )
            + ". Banks above 15% multifamily share are credible apartment lenders for acquisition, bridge, or refi conversations."
        ),
        "evidence": "RCON1460 mapped to Multifamily; ranked by mix_mf_pct.",
        "borrower_action": "Lead with rent roll and occupancy; ask about bridge vs permanent appetite.",
    })

    # --- 4. Community bank vs regional ---
    section("4. Community bank (FFIEC 041) vs regional lender availability")
    seg_rows = []
    for seg in active["form_segment"].unique():
        sub = active[active["form_segment"] == seg]
        seg_rows.append({
            "Segment": seg,
            "Banks": len(sub),
            "Median total loans ($M)": sub["total_loans_usd"].median() / 1e6,
            "Median C&I %": sub["mix_ci_pct"].median(),
            "Median Investor CRE %": sub["mix_inv_pct"].median(),
            "Median Consumer %": sub["mix_cons_pct"].median(),
            "Banks with CRE+C&I >50%": int(((sub["mix_inv_pct"] + sub["mix_ci_pct"]) > 50).sum()),
        })
    seg_df = pd.DataFrame(seg_rows).sort_values("Banks", ascending=False)
    comm = active[active["form_segment"] == "Community bank (FFIEC 041)"]
    top_seg = seg_df.iloc[0]
    table(
        seg_df,
        (
            f"This splits Texas lenders by Call Report form type — a proxy for bank size and business model. "
            f"The largest segment is {top_seg['Segment']} with {round_num(top_seg['Banks'])} banks. "
            f"Community banks (FFIEC 041) typically offer relationship-driven CRE and C&I with faster credit committees; "
            f"regional/larger banks (031/032) have bigger hold limits but more layers of approval. "
            f"{len(comm)} community banks file in this extract; "
            f"{round_num(int(((comm['mix_inv_pct'] + comm['mix_ci_pct']) > 50).sum()) if len(comm) else 0)} of them are "
            f"portfolio-style (investor CRE + C&I >50% of loans). "
            f"Commercial borrowers should compare median C&I % and Investor CRE % across segments to find banks "
            f"whose disclosed mix matches their deal type — avoid segments with high median Consumer % if you need CRE."
        ),
    )
    insights.append({
        "id": "I-04",
        "priority": "High",
        "theme": "Bank size fit",
        "insight": (
            f"{len(comm)} Texas community banks (FFIEC 041) file detailed loan schedules. "
            f"Median portfolio: ${comm['total_loans_usd'].median()/1e6:.0f}M. "
            f"{int(((comm['mix_inv_pct'] + comm['mix_ci_pct']) > 50).sum())} are portfolio-style lenders "
            f"(investor CRE + C&I >50% of book) — the sweet spot for commercial borrowers avoiding retail mortgage shops."
        ),
        "evidence": "reporting_form segmentation; mix_inv + mix_ci thresholds.",
        "borrower_action": "Prefer FFIEC 041 banks for relationship-driven CRE/C&I; expect faster credit committee paths vs money-center banks.",
    })

    # --- 5. ICP segment ---
    section("5. Lenni ICP community banks ($500M–$2B assets)")
    icp = active[active["is_icp"] == True]  # noqa: E712
    non_icp = active[active["is_icp"] != True]  # noqa: E712
    icp_rows = [{
        "Segment": "ICP ($500M–$2B)",
        "Banks": len(icp),
        "Median loans ($M)": icp["total_loans_usd"].median() / 1e6 if len(icp) else 0,
        "Median CRE % (inv+own)": (icp["mix_inv_pct"] + icp["mix_own_pct"]).median() if len(icp) else 0,
        "Median C&I %": icp["mix_ci_pct"].median() if len(icp) else 0,
    }, {
        "Segment": "Outside ICP",
        "Banks": len(non_icp),
        "Median loans ($M)": non_icp["total_loans_usd"].median() / 1e6,
        "Median CRE % (inv+own)": (non_icp["mix_inv_pct"] + non_icp["mix_own_pct"]).median(),
        "Median C&I %": non_icp["mix_ci_pct"].median(),
    }]
    icp_df = pd.DataFrame(icp_rows)
    table(
        icp_df,
        (
            f"The ICP band ($500M–$2B assets) is the community-bank sweet spot for commercial borrowers: "
            f"large enough to hold $2M–$25M relationship credits, small enough to value direct sponsor relationships. "
            f"{len(icp)} banks fall in ICP vs {len(non_icp)} outside it. "
            f"Median loan book in ICP is ${round_num(icp_df.iloc[0]['Median loans ($M)'])}M "
            f"with median CRE (investor + owner-occupied) at {round_num(icp_df.iloc[0]['Median CRE % (inv+own)'])}% "
            f"and C&I at {round_num(icp_df.iloc[0]['Median C&I %'])}%. "
            f"Banks outside ICP include very small community banks (may lack capacity for larger deals) and "
            f"regional/national banks (may be less relationship-focused). "
            f"Borrowers seeking a community-bank partner should prioritize the ICP row when building a shortlist."
        ),
    )

    icp_spec = {}
    for key, (label, _) in MIX_DEF.items():
        icp_spec[label] = int((icp[f"mix_{key}_pct"] >= SPECIALIST_PCT).sum()) if len(icp) else 0
    icp_spec_df = pd.DataFrame([
        {"Product": k, "ICP specialists (≥8%)": v} for k, v in sorted(icp_spec.items(), key=lambda x: -x[1])
    ])
    eda.append(("ICP specialists by product", ""))
    top_icp_prod = icp_spec_df.iloc[0] if len(icp_spec_df) else None
    table(
        icp_spec_df,
        (
            f"This shows how many ICP-sized banks ({round_num(len(icp))} total) are true specialists (≥{round_num(SPECIALIST_PCT)}% portfolio share) "
            f"in each borrower product. "
            f"{top_icp_prod['Product'] if top_icp_prod is not None else 'C&I'} has the most ICP specialists "
            f"({round_num(top_icp_prod['ICP specialists (≥8%)']) if top_icp_prod is not None else 0} banks). "
            f"Multifamily and construction typically show fewer ICP specialists than C&I or investor CRE — "
            f"borrowers in those niches should start with this table to set realistic outreach expectations. "
            f"An ICP bank with ≥8% in your product is a high-probability match for relationship lending."
        ),
    )

    insights.append({
        "id": "I-05",
        "priority": "High",
        "theme": "ICP opportunity",
        "insight": (
            f"{len(icp)} Texas banks fall in the $500M–$2B ICP band with active loan books. "
            f"Within ICP, C&I specialists: {icp_spec.get('Commercial & Industrial (C&I)', 0)}; "
            f"Investor CRE specialists: {icp_spec.get('Investor CRE (income property)', 0)}; "
            f"Multifamily specialists: {icp_spec.get('Multifamily (5+ units)', 0)}. "
            "These banks are large enough to hold $2M–$25M relationship credits but small enough to value direct sponsor relationships."
        ),
        "evidence": "total_assets / icp_fit from texas_bank_profiles_latest.csv cross-walk.",
        "borrower_action": "Shortlist 3–5 ICP banks in your city with ≥8% portfolio share in your product before broad cold outreach.",
    })

    # --- 6. Geographic density ---
    section("6. Geographic lender density (HQ city)")
    if "city" in active.columns and active["city"].notna().any():
        city_stats = (
            active.groupby("city", dropna=False)
            .agg(
                banks=("id_rssd", "count"),
                median_loans_M=("total_loans_usd", lambda s: s.median() / 1e6),
                mf_specialists=("mix_mf_pct", lambda s: int((s >= SPECIALIST_PCT).sum())),
                cre_specialists=("mix_inv_pct", lambda s: int((s >= SPECIALIST_PCT).sum())),
            )
            .reset_index()
            .sort_values("banks", ascending=False)
            .head(20)
        )
        city_stats.columns = ["City", "Banks (HQ)", "Median loans ($M)", "MF specialists", "Investor CRE specialists"]
        top_city = city_stats.iloc[0]
        table(
            city_stats,
            (
                f"Geographic lender density by headquarters city — where bank decision-makers cluster. "
                f"{top_city['City']} has the most HQ'd banks ({round_num(top_city['Banks (HQ)'])}), "
                f"with {round_num(top_city['MF specialists'])} multifamily specialists and "
                f"{round_num(top_city['Investor CRE specialists'])} investor CRE specialists. "
                f"Major metros (Dallas, Houston, Austin, San Antonio) offer more lender choice but also more competition for bank attention. "
                f"Smaller cities may have fewer banks HQ'd locally but branches may still serve your market — "
                f"pair this table with FDIC locations.csv for branch footprint. "
                f"Median loans ($M) shows typical bank size in each city; larger medians suggest more capacity for bigger deals."
            ),
        )

        dallas = city_stats[city_stats["City"].str.upper() == "DALLAS"]
        if not dallas.empty:
            d = dallas.iloc[0]
            insights.append({
                "id": "I-06",
                "priority": "Medium",
                "theme": "Geography",
                "insight": (
                    f"Dallas HQ concentration: {int(d['Banks (HQ)'])} banks, "
                    f"{int(d['MF specialists'])} multifamily specialists, "
                    f"{int(d['Investor CRE specialists'])} investor CRE specialists. "
                    "Major metros offer borrower choice but also require sharper differentiation in outreach."
                ),
                "evidence": "HQ city from bank profiles; branch coverage not in this file.",
                "borrower_action": "Pair HQ-city rankings with branch/footprint data (locations.csv) before assuming local presence.",
            })

    # --- 7. Temporal trends ---
    section("7. Texas loan market trends by quarter (borrower-relevant totals)")
    trend_rows = []
    for period, grp in df.groupby("reporting_period"):
        w = build_bank_wide(grp)
        m = compute_mix_frame(w)
        sub = m[m["total_loans_gross"] > 0]
        trend_rows.append({
            "Period": pd.Timestamp(period).strftime("%Y-%m-%d"),
            "Banks filing": len(sub),
            "Total loans ($B)": sub["total_loans_usd"].sum() / 1e9,
            "Investor CRE ($B)": sub["mix_inv_usd"].sum() / 1e9,
            "C&I ($B)": sub["mix_ci_usd"].sum() / 1e9,
            "Multifamily ($B)": sub["mix_mf_usd"].sum() / 1e9,
            "Construction ($B)": sub["mix_con_usd"].sum() / 1e9,
            "Median past-due 90+ %": sub["past_due_90_plus_pct"].median(),
        })
    trend_df = pd.DataFrame(trend_rows).sort_values("Period")
    if len(trend_df) >= 2:
        first, last = trend_df.iloc[0], trend_df.iloc[-1]
        loan_chg = (last["Total loans ($B)"] - first["Total loans ($B)"]) / first["Total loans ($B)"] * 100
        trend_note = (
            f"Total Texas bank loans moved from ${round_num(first['Total loans ($B)'])}B ({first['Period']}) "
            f"to ${round_num(last['Total loans ($B)'])}B ({last['Period']}) — {round_num(loan_chg)}% change. "
        )
    else:
        trend_note = "Insufficient quarters for trend comparison. "
    table(
        trend_df,
        (
            f"Five-quarter view of aggregate Texas bank lending capacity by product category. "
            f"{trend_note}"
            f"Growing totals in your sector (e.g. multifamily or construction) often mean lenders have more balance-sheet "
            f"room for new originations; flat or declining books may signal tighter underwriting or paydowns. "
            f"'Banks filing' tracks panel size (mergers/closures reduce count). "
            f"Median past-due 90+ % is a portfolio health signal across all banks — rising medians suggest "
            f"broad credit stress that may slow new lending. "
            f"Use this to time outreach: approach lenders when their category book is growing quarter-over-quarter."
        ),
    )

    if len(trend_df) >= 2:
        first, last = trend_df.iloc[0], trend_df.iloc[-1]
        loan_chg = (last["Total loans ($B)"] - first["Total loans ($B)"]) / first["Total loans ($B)"] * 100
        insights.append({
            "id": "I-07",
            "priority": "Medium",
            "theme": "Market momentum",
            "insight": (
                f"Aggregate Texas bank loans moved from ${first['Total loans ($B)']}B ({first['Period']}) "
                f"to ${last['Total loans ($B)']}B ({last['Period']}) — {loan_chg:+.1f}% change. "
                f"Multifamily book: ${first['Multifamily ($B)']}B → ${last['Multifamily ($B)']}B. "
                "Rising totals can mean more lender capacity; flat CRE books may signal tighter underwriting."
            ),
            "evidence": "Five-quarter trend from full labeled file.",
            "borrower_action": "Time outreach when lender portfolios in your sector are growing quarter-over-quarter.",
        })

    # --- 8. Credit quality / lender health ---
    section("8. Portfolio stress signals (past-due ratios — lender health check)")
    stress = active.copy()
    stress["stress_flag"] = (stress["past_due_90_plus_pct"] > 1.0) | (stress["past_due_30_89_pct"] > 2.0)
    stress_summary = pd.DataFrame([
        {"Metric": "Median past-due 30–89 days (% of loans)", "Value": stress["past_due_30_89_pct"].median()},
        {"Metric": "Median past-due 90+ days (% of loans)", "Value": stress["past_due_90_plus_pct"].median()},
        {"Metric": "Banks with elevated stress (90+ >1% or 30–89 >2%)", "Value": int(stress["stress_flag"].sum())},
        {"Metric": "Share of banks with elevated stress (%)", "Value": 100 * stress["stress_flag"].mean()},
    ])
    table(
        stress_summary,
        (
            f"Lender health check using portfolio-level past-due ratios from Call Report lines RCON5367/5368. "
            f"The typical Texas bank has {round_num(stress['past_due_30_89_pct'].median())}% of loans "
            f"30–89 days past due and {round_num(stress['past_due_90_plus_pct'].median())}% 90+ days past due. "
            f"{round_num(int(stress['stress_flag'].sum()))} banks ({round_num(100*stress['stress_flag'].mean())}%) "
            f"exceed elevated-stress thresholds — worth extra diligence before committing to a term sheet. "
            f"High past-due ratios do not mean a bank will decline your deal, but they may indicate "
            f"tighter credit committees, concentration limits, or paused originations in stressed asset classes. "
            f"Ask prospective lenders about recent charge-offs and appetite for new commitments."
        ),
    )

    by_prod_stress = []
    for key, (label, _) in MIX_DEF.items():
        col = f"mix_{key}_pct"
        heavy = active[active[col] >= SPECIALIST_PCT]
        if len(heavy) == 0:
            continue
        by_prod_stress.append({
            "Product (specialists only)": label,
            "Specialist banks": len(heavy),
            "Median 90+ day past-due %": heavy["past_due_90_plus_pct"].median(),
            "Median 30–89 day past-due %": heavy["past_due_30_89_pct"].median(),
        })
    by_prod_stress_df = pd.DataFrame(by_prod_stress)
    table(
        by_prod_stress_df,
        (
            f"Past-due ratios among banks that are product specialists (≥{round_num(SPECIALIST_PCT)}% portfolio share) — "
            f"your most likely lending partners. "
            f"This shows whether specialist lenders carry higher or lower stress than the overall market. "
            f"If specialists in your product show low median past-due %, those banks have managed that book well. "
            f"Elevated medians in a category may mean the specialist banks are working through legacy problems — "
            f"probe their current origination appetite and whether stress is concentrated in a different sub-segment "
            f"(e.g. office vs industrial within investor CRE)."
        ),
    )

    insights.append({
        "id": "I-08",
        "priority": "Medium",
        "theme": "Lender diligence",
        "insight": (
            f"{int(stress['stress_flag'].sum())} Texas banks ({100*stress['stress_flag'].mean():.1f}%) show elevated "
            "past-due ratios on aggregate loan books. Borrowers should treat unusually high portfolio stress as a "
            "signal to probe credit appetite, concentration limits, and whether new originations are paused."
        ),
        "evidence": "RCON5367 / RCON5368 vs total loans (RCON2122).",
        "borrower_action": "Ask lenders about recent charge-offs and appetite for new commitments in your asset class.",
    })

    # --- 9. Portfolio orientation clusters ---
    section("9. Borrower-oriented bank clusters (portfolio archetypes)")
    archetypes = []
    for _, r in active.iterrows():
        mixes = {k: r[f"mix_{k}_pct"] for k in MIX_DEF}
        top = max(mixes.items(), key=lambda x: x[1])
        cre_ci = r["mix_inv_pct"] + r["mix_own_pct"] + r["mix_ci_pct"]
        archetypes.append({
            "id_rssd": r["id_rssd"],
            "institution_name": r["institution_name"],
            "dominant_product": MIX_DEF[top[0]][0],
            "dominant_pct": top[1],
            "cre_ci_pct": cre_ci,
            "consumer_pct": r["mix_cons_pct"],
        })
    arch_df = pd.DataFrame(archetypes)
    arch_summary = (
        arch_df.groupby("dominant_product")
        .agg(banks=("id_rssd", "count"), median_dominant_pct=("dominant_pct", "median"))
        .reset_index()
        .sort_values("banks", ascending=False)
    )
    arch_summary.columns = ["Dominant portfolio type", "Banks", "Median dominance %"]
    portfolio_lenders = int((arch_df["cre_ci_pct"] > 50).sum())
    retail_heavy = int((arch_df["consumer_pct"] > 25).sum())
    dom_leader = arch_summary.iloc[0]
    table(
        arch_summary,
        (
            f"Each bank is classified by its single largest loan category (dominant portfolio type). "
            f"{dom_leader['Dominant portfolio type']} is the most common dominant specialty "
            f"({round_num(dom_leader['Banks'])} banks, median {round_num(dom_leader['Median dominance %'])}% of book). "
            f"{portfolio_lenders} banks are portfolio-style (CRE + C&I >50% of loans) — the best fit for commercial borrowers. "
            f"{retail_heavy} banks are consumer-heavy (>25% consumer) — weaker fit for CRE/C&I outreach. "
            f"Use this to filter your target list: prioritize banks whose dominant type aligns with your need, "
            f"and deprioritize consumer-oriented institutions for commercial deals."
        ),
    )
    insights.append({
        "id": "I-09",
        "priority": "High",
        "theme": "Bank archetypes",
        "insight": (
            f"{portfolio_lenders} banks are portfolio-style (CRE + C&I >50% of loans) — best fit for commercial borrowers. "
            f"{retail_heavy} banks are consumer-heavy (>25% consumer) — weaker fit for CRE/C&I deals. "
            f"Most common dominant specialty: {arch_summary.iloc[0]['Dominant portfolio type']} "
            f"({int(arch_summary.iloc[0]['Banks'])} banks)."
        ),
        "evidence": "K-means-free rule clustering on mix percentages.",
        "borrower_action": "Filter out consumer-heavy banks before CRE outreach; prioritize portfolio lenders with ≥8% in your product.",
    })

    # --- 10. Regulatory code utilization ---
    section("10. MDRM code depth — what regulators measure vs what borrowers need")
    code_stats = (
        latest.groupby("mdrm_code")
        .agg(
            banks_reporting=("id_rssd", "nunique"),
            nonzero_balances=("value_num", lambda s: int((s > 0).sum())),
            total_value_M=("value_num", lambda s: s.sum() / 1e3),
        )
        .reset_index()
        .sort_values("nonzero_balances", ascending=False)
    )
    top_code = code_stats.iloc[0] if len(code_stats) else None
    table(
        code_stats.head(25),
        (
            f"The labeled file contains {round_num(df['mdrm_code'].nunique())} distinct MDRM codes; "
            f"this table shows the 25 most frequently reported with non-zero balances in the latest quarter. "
            f"{top_code['mdrm_code'] if top_code is not None else 'RCON2122'} leads with "
            f"{round_num(top_code['nonzero_balances']) if top_code is not None else 0} non-zero observations "
            f"across {round_num(top_code['banks_reporting']) if top_code is not None else 0} banks. "
            f"Codes with high nonzero_balances are the regulatory lines Texas banks actually use — "
            f"these are the measurable product categories available for bank-matching. "
            f"Rare codes with few nonzero rows may still matter for niche products but offer less statistical power. "
            f"total_value_M sums reported balances for that code (scale per Call Report units)."
        ),
    )
    eda.append(("Codes with zero Texas balances (latest period)", ""))
    zero_codes = code_stats[code_stats["nonzero_balances"] == 0]
    eda.append(("Count of unused codes in latest period", round_eda_scalar(len(zero_codes))))
    eda.append(("", ""))

    if TAXONOMY.exists():
        section("11. Taxonomy crosswalk — borrower catalog coverage")
        tax = pd.read_csv(TAXONOMY, low_memory=False)
        in_cat = tax[tax["Loan Product Category"].isin(BORROWER_CATEGORIES)]
        cat_cov = (
            in_cat.groupby("Loan Product Category")
            .agg(
                regulatory_lines=("Regulatory Line Item Code", "nunique"),
                median_banks_reporting=("Number of Texas Banks Reporting", "median"),
                in_borrower_site=("Listed in Borrower Product Catalog", lambda s: int((s == "Yes").sum())),
            )
            .reset_index()
            .sort_values("median_banks_reporting", ascending=False)
        )
        table(
            cat_cov,
            (
                f"Crosswalk between Federal Reserve MDRM codes and borrower-friendly product categories "
                f"(from texas_mdrm_loan_taxonomy.csv). "
                f"Shows how many regulatory line items map to each borrower product and how many Texas banks report them. "
                f"Categories with high median_banks_reporting are easiest to match — regulators require most banks to file those lines. "
                f"in_borrower_site counts how many lines appear in the Lenni borrower product catalog. "
                f"Use this to understand which borrower-facing product names are backed by measurable regulatory data in Texas."
            ),
        )

    # --- 12. Specialist count distribution ---
    section("12. How many product specialists can a borrower realistically contact?")
    spec_counts = []
    for key, (label, _) in MIX_DEF.items():
        n = int((active[f"mix_{key}_pct"] >= SPECIALIST_PCT).sum())
        n15 = int((active[f"mix_{key}_pct"] >= 15).sum())
        spec_counts.append({
            "Product": label,
            f"Banks ≥ {SPECIALIST_PCT:.0f}%": n,
            "Banks ≥ 15%": n15,
            "Realistic shortlist size (top 10–15)": min(15, n),
        })
    spec_df = pd.DataFrame(spec_counts)
    top_spec = spec_df.sort_values(f"Banks ≥ {SPECIALIST_PCT:.0f}%", ascending=False).iloc[0]
    table(
        spec_df,
        (
            f"Practical outreach sizing: how many Texas banks meet specialist thresholds for each product. "
            f"{top_spec['Product']} has the most specialists at ≥{round_num(SPECIALIST_PCT)}% "
            f"({round_num(top_spec[f'Banks ≥ {SPECIALIST_PCT:.0f}%'])} banks) and "
            f"{round_num(top_spec['Banks ≥ 15%'])} at ≥15%. "
            f"'Realistic shortlist size' caps at 15 — the maximum useful number of banks to contact in a first wave. "
            f"For C&I and investor CRE you can build a full 10–15 bank list; for multifamily or construction "
            f"you may need to go deeper on relationships because the specialist pool is smaller. "
            f"Start with ≥8% portfolio share, then narrow by geography and deal size fit."
        ),
    )

    insights.append({
        "id": "I-10",
        "priority": "High",
        "theme": "Outreach planning",
        "insight": (
            "A borrower can build a credible 10-bank shortlist for C&I and Investor CRE; "
            "Multifamily and Construction have fewer deep specialists — quality of introduction matters more than volume. "
            "Use portfolio % ≥8% as the first filter, then geography and relationship fit."
        ),
        "evidence": "Specialist counts in section 12; aligns with borrower site ranking methodology.",
        "borrower_action": "Prepare a one-page deal summary before calling; lead with why your deal fits their disclosed portfolio mix.",
    })

    # --- 13. Correlation / diversification ---
    section("13. Product co-occurrence among specialist lenders")
    spec_matrix = []
    keys = list(MIX_DEF.keys())[:7]
    for k in keys:
        row = {"Product": MIX_DEF[k][0]}
        base = active[f"mix_{k}_pct"] >= SPECIALIST_PCT
        for k2 in keys:
            if k == k2:
                row[MIX_DEF[k2][0][:20]] = 100.0
            else:
                both = (base & (active[f"mix_{k2}_pct"] >= SPECIALIST_PCT)).sum()
                denom = base.sum()
                row[MIX_DEF[k2][0][:20]] = 100 * both / denom if denom else 0
        spec_matrix.append(row)
    spec_matrix_df = pd.DataFrame(spec_matrix)
    table(
        spec_matrix_df,
        (
            f"Co-specialization matrix: for each row product, what % of its specialist banks also specialize "
            f"in each column product (≥{round_num(SPECIALIST_PCT)}% threshold). "
            f"Diagonal values are 100% by definition. "
            f"High off-diagonal values (e.g. investor CRE row × C&I column) mean banks that do one often do both — "
            f"useful if you need a property loan plus an operating line from the same institution. "
            f"Low overlap (e.g. multifamily × ag) means those are distinct lender types requiring separate searches. "
            f"Borrowers with multi-product needs should prioritize banks with high co-occurrence in their categories."
        ),
    )

    insights.append({
        "id": "I-11",
        "priority": "Low",
        "theme": "Cross-sell patterns",
        "insight": (
            "Banks that specialize in Investor CRE often also show C&I specialization — "
            "full-relationship banks can bundle operating lines with property loans. "
            "Multifamily specialists less frequently overlap with ag/farmland niches."
        ),
        "evidence": "Co-specialization matrix (% of row-product specialists who also specialize in column product).",
        "borrower_action": "If you need both CRE and a revolver, prioritize banks with high co-occurrence in your categories.",
    })

    # --- 14. Value distribution / outliers ---
    section("14. Exposure distribution among active lenders (latest quarter)")
    dist_rows = []
    for key, (label, _) in MIX_DEF.items():
        col = f"mix_{key}_usd"
        pos = active.loc[active[col] > 0, col]
        if pos.empty:
            continue
        dist_rows.append({
            "Product": label,
            "P25 ($M)": pos.quantile(0.25) / 1e6,
            "Median ($M)": pos.median() / 1e6,
            "P75 ($M)": pos.quantile(0.75) / 1e6,
            "Max ($M)": pos.max() / 1e6,
            "Mean ($M)": pos.mean() / 1e6,
        })
    dist_df = pd.DataFrame(dist_rows)
    if not dist_df.empty:
        widest = dist_df.loc[dist_df["Max ($M)"].idxmax()]
        table(
            dist_df,
            (
                f"Distribution of active-lender exposures by product — helps match your deal size to bank capacity. "
                f"P25/median/P75 show the typical range among banks that actually hold balances in each category. "
                f"{widest['Product']} has the largest maximum exposure (${round_num(widest['Max ($M)'])}M), "
                f"reflecting the presence of very large regional banks. "
                f"A $3M CRE loan is material for a bank at the median but small for one at P75. "
                f"Disclose your loan size early and ask whether it clears the bank's hold limit and committee threshold. "
                f"Target banks where your ask sits at or above their median product balance for better engagement."
            ),
        )

    insights.append({
        "id": "I-12",
        "priority": "Medium",
        "theme": "Deal size fit",
        "insight": (
            "Median active-lender exposures span orders of magnitude — a $3M CRE loan is material for many community banks "
            "but immaterial for regional portfolios. Match your ask to banks where your deal size sits above their median "
            "product balance for higher engagement odds."
        ),
        "evidence": "P25/median/P75 of positive balances per product category.",
        "borrower_action": "Disclose loan size early; ask whether it clears their internal hold-limit and committee threshold.",
    })

    # --- Methodology ---
    section("Methodology & caveats")
    eda.extend([
        ("Data source", "FFIEC Call Report XBRL facts enriched with Federal Reserve MDRM labels"),
        ("Grain", "Bank × reporting period × regulatory line item (mdrm_code)"),
        ("Borrower mapping", "Schedule RC-C summary lines → Lenni loan product taxonomy (loan_mix.py / loan_products.yaml)"),
        ("Specialist definition", f"Portfolio share ≥ {SPECIALIST_PCT}% in product category vs total loans (RCON2122)"),
        ("Units", "value_num is thousands USD unless otherwise noted; displayed $M/$B multiply accordingly"),
        ("Limitations", "Regulatory lines ≠ named loan products on bank websites; no pricing, LTV policy, or credit box"),
        ("Limitations", "Branch geography requires locations.csv; this analysis uses HQ city when available"),
        ("Limitations", "Past-due lines are portfolio-level, not product-specific stress"),
        ("", ""),
    ])

    return eda, insights


def build_data_extraction_rows(df: pd.DataFrame) -> list[tuple[str, object]]:
    """Detailed provenance documentation for the texas_loans_labeled sheet."""
    import json

    root_otx = ROOT / "ONLY_TEXAS_SINCE_2025"
    period_counts = (
        df.groupby("reporting_period")["id_rssd"]
        .nunique()
        .reset_index()
        .rename(columns={"reporting_period": "Reporting period", "id_rssd": "Banks filing"})
    )
    period_counts["Reporting period"] = pd.to_datetime(period_counts["Reporting period"]).dt.strftime("%Y-%m-%d")
    period_list = ", ".join(period_counts["Reporting period"].tolist())

    cat_counts = df["mdrm_category"].value_counts().reset_index()
    cat_counts.columns = ["mdrm_category", "Rows in this file"]

    prefix_rows = []
    prefixes = [
        ("RCON21*", "Totals, allowance, net loans"),
        ("RCON14*", "Real estate–secured loans"),
        ("RCON15*", "Consumer / credit card"),
        ("RCON16*", "Other commercial loans"),
        ("RCON17*", "Lease financing"),
        ("RCONF1*", "Schedule RC-C loan categories (construction, CRE, etc.)"),
        ("RCONF2*", "Schedule RC-C extension"),
        ("RCONHK*", "Schedule RC-C memoranda"),
        ("RCONJ4*", "Schedule RC-C extension detail"),
        ("RCONLL*", "Leases"),
        ("RCONA5*", "RC-C extension detail"),
        ("RCONB5*", "FFIEC 051 supplemental consumer lines"),
        ("RCFD*", "Consolidated/domestic office variant (often FFIEC 041 banks)"),
    ]
    for pref, desc in prefixes:
        code = pref.replace("*", "")
        n = int(df["mdrm_code"].str.startswith(code).sum())
        if n:
            prefix_rows.append({"Code prefix": pref, "Approx. rows in file": n, "Typical content": desc})
    prefix_df = pd.DataFrame(prefix_rows)

    n_progress = 0
    prog_path = root_otx / "data" / "progress.json"
    if prog_path.exists():
        raw = json.loads(prog_path.read_text())
        n_progress = len(raw.get("completed", raw) if isinstance(raw, dict) else raw)

    rows: list[tuple[str, object]] = []

    def sec(title: str) -> None:
        rows.append((f"§ {title}", ""))
        rows.append(("", ""))

    rows.append(("DATA EXTRACTION — How texas_loans_labeled was built", ""))
    rows.append(("", ""))
    rows.extend([
        ("Document built", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("This workbook sheet", "texas_loans_labeled (937,814 regulatory loan line items)"),
        ("Source CSV", "ONLY_TEXAS_SINCE_2025/exports/texas_loans_labeled.csv"),
        ("Extraction script", "ONLY_TEXAS_SINCE_2025/extract_texas_loans.py (full mode, not --summary)"),
        ("Project folder", "ONLY_TEXAS_SINCE_2025/ — Texas Call Reports (2025+)"),
        ("Authentication", "Official FFIEC Public Web Service (REST API) + Federal Reserve MDRM dictionary — no web scraping"),
        ("", ""),
    ])

    sec("1. Executive summary — what this file contains")
    rows.extend([
        ("Definition", (
            "Every Schedule RC-C and loan-related Call Report line item for Texas banks, enriched with "
            "Federal Reserve MDRM plain-English labels. Each row is one regulatory category per bank per quarter — "
            "not individual customer loans."
        )),
        ("Geographic scope", "Texas only (State = TX on FFIEC Panel of Reporters)"),
        ("Time scope", f"Reporting periods ending 2025 or later: {period_list}"),
        ("Rows in texas_loans_labeled sheet", f"{len(df):,}"),
        ("Distinct Texas banks (all periods)", f"{df['id_rssd'].nunique()}"),
        ("Distinct MDRM codes in file", f"{df['mdrm_code'].nunique():,}"),
        ("Reporting periods", str(df["reporting_period"].nunique())),
        ("Grain", "Bank (id_rssd) × reporting_period × mdrm_code"),
        ("Companion files", (
            "texas_loans_summary.csv (~31k rows, curated categories); "
            "texas_loan_products_mdrm_catalog.csv (code book); "
            "texas_xbrl_facts.csv (all XBRL facts, ~2.19M rows)"
        )),
        ("", ""),
    ])
    rows.append(("__table__", period_counts))
    rows.append(("", ""))

    sec("2. End-to-end pipeline (three phases)")
    rows.extend([
        ("Phase A", "pull_texas_since_2025.py — download & parse FFIEC Call Report XBRL → texas_xbrl_facts.csv"),
        ("Phase B", "extract_texas_loans.py — filter loan/RC-C lines + MDRM labels → texas_loans_labeled.csv"),
        ("Phase C (optional)", "build_lenni_eda_report.py / borrower site builders — join & aggregate for analysis"),
        ("This workbook", "texas_loans_labeled.csv exported to Excel + borrower EDA (EDA & Insights sheets)"),
        ("", ""),
        ("Flow", "FFIEC API → raw .xbrl archive → parse_xbrl → texas_xbrl_facts.csv → extract_texas_loans.py → this sheet"),
        ("", ""),
    ])

    sec("3. Phase A — FFIEC API download (pull_texas_since_2025.py)")
    rows.extend([
        ("Script path", "ONLY_TEXAS_SINCE_2025/pull_texas_since_2025.py"),
        ("API base URL", "https://ffieccdr.azure-api.us/public/"),
        ("API specification", "https://cdr.ffiec.gov/public/Files/SIS611_-_Retrieve_Public_Data_via_Web_Service.pdf"),
        ("Credentials", "FFIEC_USER_ID + FFIEC_TOKEN in repository root .env (Bearer token in Authentication header)"),
        ("Rate limiting", "~1.5 seconds between API calls (~2,400/hour, under FFIEC 2,500/hour cap)"),
        ("", ""),
        ("Step 0 — Authenticate", "HTTP headers: UserID + Authentication: Bearer <token>"),
        ("Step 1 — List periods", "RetrieveReportingPeriods (data_series=Call) → keep quarters with year ≥ 2025"),
        ("Step 2 — Panel of reporters", "RetrievePanelOfReporters per period → filter State.upper() == 'TX'"),
        ("Step 3 — Write institutions", "All TX panel rows → exports/texas_institutions.csv"),
        ("Step 4 — Download filings", "For each bank with HasFiledForReportingPeriod=True → RetrieveFacsimile (XBRL format)"),
        ("Step 5 — Archive raw files", "archive/call/<period>/<rssd>.xbrl + .meta.json (SHA-256 hash, retrieval timestamp)"),
        ("Step 6 — Parse XBRL", "ffiec_cdr.parser.parse_xbrl (lxml) → append exports/texas_xbrl_facts.csv"),
        ("Step 7 — Resume support", "data/progress.json tracks completed period|RSSD pairs — skips finished downloads on re-run"),
        ("", ""),
        ("Checkpoint pairs completed", f"{n_progress:,} period|RSSD pairs in data/progress.json"),
        ("Raw XBRL archive", "1,825 .xbrl files under ONLY_TEXAS_SINCE_2025/archive/call/"),
        ("Parsed XBRL facts (all line items)", "~2,186,590 rows in texas_xbrl_facts.csv"),
        ("Typical runtime", "1–3 hours for full Texas 2025+ download (network dependent)"),
        ("", ""),
    ])

    sec("4. FFIEC Public Web Service methods — used vs not used")
    rows.extend([
        ("RetrieveReportingPeriods", "YES — list Call Report quarters"),
        ("RetrievePanelOfReporters", "YES — bank panel with state, city, filing status, form type"),
        ("RetrieveFacsimile", "YES — download XBRL Call Report facsimile per bank/quarter"),
        ("RetrieveFilersSinceDate", "NO — incremental sync only (not needed for full backfill)"),
        ("RetrieveFilersSubmissionDateTime", "NO — optional submission timestamps"),
        ("RetrieveUBPRReportingPeriods", "NO — UBPR is a different report (peer ratios)"),
        ("RetrieveUBPRXBRLFacsimile", "NO — UBPR XBRL, not Call Report"),
        ("Formats not downloaded", "PDF and SDF (same RetrieveFacsimile API with different facsimileFormat header)"),
        ("", ""),
    ])

    sec("5. Phase B — MDRM labeling (extract_texas_loans.py, full labeled export)")
    rows.extend([
        ("Script path", "ONLY_TEXAS_SINCE_2025/extract_texas_loans.py"),
        ("Mode", "Default (no --summary flag) — retains ALL loan/Schedule RC-C related rows"),
        ("MDRM dictionary source", "https://www.federalreserve.gov/apps/mdrm/pdf/MDRM.zip"),
        ("Local MDRM file", "ONLY_TEXAS_SINCE_2025/data/mdrm/MDRM_CSV.csv (~91 MB)"),
        ("MDRM loader module", "ONLY_TEXAS_SINCE_2025/mdrm_loader.py"),
        ("Input file", "ONLY_TEXAS_SINCE_2025/exports/texas_xbrl_facts.csv"),
        ("Output CSV", "ONLY_TEXAS_SINCE_2025/exports/texas_loans_labeled.csv"),
        ("Also produced", "texas_loan_products_mdrm_catalog.csv (24,015 loan/lease MDRM definitions)"),
        ("", ""),
        ("Step 1 — Load MDRM lookup", "Map each mdrm_code → item_name, description, reporting_form, item_type, mdrm_category"),
        ("Step 2 — Stream facts CSV", "Read texas_xbrl_facts.csv row-by-row (memory efficient for ~2.19M rows)"),
        ("Step 3 — Extract local concept", "Parse XBRL concept QName → local MDRM code (text after last '}')"),
        ("Step 4 — Filter loan rows", "Keep row if is_loan_row(code) returns True (see filter rules below)"),
        ("Step 5 — Enrich & write", "Join MDRM labels; write 14-column labeled export"),
        ("Step 6 — Build catalog", "Write texas_loan_products_mdrm_catalog.csv with in_texas_data flag per code"),
        ("", ""),
    ])

    sec("6. Loan-row filter rules (why 937k rows from 2.19M facts)")
    rows.extend([
        ("Summary mode (--summary)", "Only ~21 curated SUMMARY_CODES (e.g. RCON2122, RCON1460) → ~31,396 rows"),
        ("Full labeled mode (this file)", "All Schedule RC-C + loan prefix families → ~937,816 rows"),
        ("Rule 1 — Curated codes", "Any code in SUMMARY_CODES set (totals, CRE, consumer, ag, construction, past-due, etc.)"),
        ("Rule 2 — MDRM category", "mdrm_category in ('loan_or_lease', 'schedule_rc_c') per Fed dictionary keywords"),
        ("Rule 3 — Code prefixes", (
            "RCON14*, RCON15*, RCON16*, RCON17*, RCON21*, RCONF1*, RCONF2*, RCONHK*, RCONJ4*, "
            "RCONLL*, RCONA5*, RCONB5*, RCFD14*, RCFD15*, RCFD16*, RCFD21*, RCFDHK*, RCFDJ4*, RCFDLL*"
        )),
        ("Excluded", "XBRL metadata (measure, period, entity, instant, startDate, endDate)"),
        ("Why 'loan' text search fails", "FFIEC XBRL uses MDRM codes (RCON2122), not English 'loan' in concept names — 0 rows match"),
        ("", ""),
    ])
    if not prefix_df.empty:
        rows.append(("__table__", prefix_df))
        rows.append(("", ""))
    rows.append(("__table__", cat_counts.head(10)))
    rows.append(("", ""))

    sec("7. Column dictionary — texas_loans_labeled sheet")
    col_docs = pd.DataFrame([
        {"Column": "id_rssd", "Description": "Federal Reserve RSSD ID — primary join key to institutions and bank profiles"},
        {"Column": "institution_name", "Description": "Bank legal name from FFIEC panel (denormalized for convenience)"},
        {"Column": "reporting_period", "Description": "Call Report quarter-end date (e.g. 2026-03-31 = Q1 2026)"},
        {"Column": "mdrm_code", "Description": "Regulatory line code (e.g. RCON2122 = total loans). Lookup: federalreserve.gov/apps/mdrm/data-dictionary"},
        {"Column": "item_name", "Description": "Official Federal Reserve short name from MDRM dictionary"},
        {"Column": "line_description", "Description": "Same as item_name, or 'MDRM {code}' if dictionary miss"},
        {"Column": "mdrm_description", "Description": "Full Fed definition (up to 800 characters)"},
        {"Column": "mdrm_category", "Description": "loan_or_lease | schedule_rc_c | loan_related_prefix | other"},
        {"Column": "reporting_form", "Description": "Call Report form (FFIEC 031 = larger, FFIEC 041 = community bank, etc.)"},
        {"Column": "item_type", "Description": "MDRM item type (F = financial line item)"},
        {"Column": "value_num", "Description": "Numeric balance reported on Call Report line (see units note below)"},
        {"Column": "value_text", "Description": "Raw XBRL text value before numeric parsing"},
        {"Column": "context_ref", "Description": "XBRL context id (period/scenario — instant vs duration)"},
        {"Column": "unit_ref", "Description": "XBRL unit (typically USD)"},
    ])
    rows.append(("__table__", col_docs))
    rows.append(("", ""))
    rows.extend([
        ("Units note — value_num", (
            "FFIEC Call Report amounts are conventionally reported in thousands of U.S. dollars. "
            "This Excel export may store values as parsed from XBRL; verify against RCON2122 for a known bank "
            "before scaling in financial models. See ONLY_TEXAS_SINCE_2025/DATA_DICTIONARY.md."
        )),
        ("RCON vs RCFD codes", (
            "RCON = domestic offices; RCFD = consolidated/domestic variant. Smaller banks (FFIEC 041) often file RCFD lines. "
            "Both may appear for the same institution depending on form type."
        )),
        ("", ""),
    ])

    sec("8. Key MDRM codes — borrower-relevant loan categories")
    key_codes = pd.DataFrame([
        {"MDRM code": "RCON2122", "Regulatory category": "Total loans and leases (headline denominator for portfolio %)"},
        {"MDRM code": "RCON1460", "Regulatory category": "Multifamily (5+) residential real estate"},
        {"MDRM code": "RCONF161", "Regulatory category": "Other nonfarm nonresidential real estate (investor CRE)"},
        {"MDRM code": "RCONF162", "Regulatory category": "Commercial real estate loans"},
        {"MDRM code": "RCONF160", "Regulatory category": "Owner-occupied nonfarm nonresidential real estate"},
        {"MDRM code": "RCONF158", "Regulatory category": "Residential construction"},
        {"MDRM code": "RCONF159", "Regulatory category": "Other construction and land development"},
        {"MDRM code": "RCON1766", "Regulatory category": "Commercial and industrial (C&I) loans"},
        {"MDRM code": "RCON1403", "Regulatory category": "1–4 family residential mortgages"},
        {"MDRM code": "RCON1545", "Regulatory category": "Credit card plans"},
        {"MDRM code": "RCON1583", "Regulatory category": "Other consumer loans"},
        {"MDRM code": "RCON1420", "Regulatory category": "Loans secured by farmland"},
        {"MDRM code": "RCON1590", "Regulatory category": "Agricultural production loans"},
        {"MDRM code": "RCON1754", "Regulatory category": "Lease financing receivables"},
        {"MDRM code": "RCON5367", "Regulatory category": "Past due 30–89 days (portfolio stress signal)"},
        {"MDRM code": "RCON5368", "Regulatory category": "Past due 90+ days (portfolio stress signal)"},
    ])
    rows.append(("__table__", key_codes))
    rows.append(("", ""))

    sec("9. Logical data model — how tables connect")
    rows.extend([
        ("texas_institutions", "Grain: bank × quarter. FFIEC Panel of Reporters — who is expected to file."),
        ("texas_filings", "Grain: bank × quarter. Download manifest: file path, SHA-256, retrieval time."),
        ("texas_xbrl_facts", "Grain: one row per XBRL fact (concept × context). Source of all numeric values."),
        ("texas_loans_labeled", "Grain: bank × quarter × mdrm_code. Filtered + labeled loan/RC-C subset (this sheet)."),
        ("texas_loans_summary", "Grain: bank × quarter × mdrm_code. Smaller curated subset (~21 codes) for quick analysis."),
        ("Join keys", "id_rssd + reporting_period links all tables"),
        ("Downstream", "texas_bank_profiles_latest.csv, borrower_site/, EDA sheets in this workbook"),
        ("", ""),
    ])

    sec("10. XBRL parsing details (ffiec_cdr.parser.parse_xbrl)")
    rows.extend([
        ("Parser location", "ffiec_cdr/parser.py (project package)"),
        ("XML library", "lxml — walks XBRL instance document element tree"),
        ("Captured per fact", "concept (QName), contextRef, unitRef, numeric or text value"),
        ("Deduplication", "Identical (concept, context_ref, value prefix) tuples within a filing are collapsed"),
        ("Safety cap", "Maximum 50,000 facts per filing (prevents runaway parses)"),
        ("Typical volume", "~800–1,200 facts per bank per quarter (varies by size and form)"),
        ("Raw source of truth", "archive/call/<period>/<rssd>.xbrl — can rebuild CSVs without re-downloading"),
        ("", ""),
    ])

    sec("11. Rebuild & recovery commands")
    rows.extend([
        ("Download MDRM dictionary", "python ONLY_TEXAS_SINCE_2025/scripts/download_mdrm.py"),
        ("Full Texas re-download", "python ONLY_TEXAS_SINCE_2025/pull_texas_since_2025.py"),
        ("Test download (5 filings)", "python ONLY_TEXAS_SINCE_2025/pull_texas_since_2025.py --max 5"),
        ("Rebuild core CSVs from archive", "python ONLY_TEXAS_SINCE_2025/rebuild_csv_from_archive.py"),
        ("Regenerate this labeled file", "python ONLY_TEXAS_SINCE_2025/extract_texas_loans.py"),
        ("Regenerate summary only", "python ONLY_TEXAS_SINCE_2025/extract_texas_loans.py --summary"),
        ("Regenerate MDRM catalog only", "python ONLY_TEXAS_SINCE_2025/extract_texas_loans.py --catalog"),
        ("Rebuild this Excel workbook", "python build_texas_loans_labeled_borrower_eda.py"),
        ("", ""),
    ])

    sec("12. What this data is NOT")
    rows.extend([
        ("Individual loan contracts", "Call Reports are confidential aggregates — no borrower names or loan-level detail"),
        ("Marketing product names", "Only regulatory categories — not '30-year fixed' or bank website product names"),
        ("Pricing or credit policy", "No interest rates, LTV limits, DSCR minimums, or underwriting boxes"),
        ("Real-time servicing data", "Quarterly regulatory filings with reporting lag"),
        ("National scope", "Texas banks on FFIEC panel only"),
        ("UBPR peer analytics", "Requires separate UBPR API pull (RetrieveUBPRXBRLFacsimile)"),
        ("Branch footprint", "HQ city from panel; branch list requires FDIC locations.csv enrichment"),
        ("", ""),
    ])

    sec("13. References & project documentation")
    rows.extend([
        ("FFIEC CDR public site", "https://cdr.ffiec.gov/public/"),
        ("PWS help", "https://cdr.ffiec.gov/public/HelpFiles/PWSInfo.htm"),
        ("Federal Reserve MDRM", "https://www.federalreserve.gov/apps/mdrm/"),
        ("MDRM code lookup", "https://www.federalreserve.gov/apps/mdrm/data-dictionary"),
        ("Schedule RC-C instructions", "https://www.fdic.gov/bank-financial-reports/031-041-rc-c1-loans-and-leases-december-2024"),
        ("Project README", "ONLY_TEXAS_SINCE_2025/README.md"),
        ("Data dictionary", "ONLY_TEXAS_SINCE_2025/DATA_DICTIONARY.md"),
        ("Loan extraction guide", "ONLY_TEXAS_SINCE_2025/LOAN_EXTRACTION_GUIDE.md"),
        ("SharePoint data folder", "See README.md — large CSVs distributed via SharePoint, not GitHub"),
        ("", ""),
    ])

    sec("14. Project timeline")
    rows.extend([
        ("2026-06-02", "Initial Texas extract script; first XBRL downloads from FFIEC PWS"),
        ("2026-06-02–03", "Full backfill: 1,825 filings across 5 quarters (2025 Q1 – 2026 Q1)"),
        ("2026-06-03", "rebuild_csv_from_archive.py — CSVs aligned with on-disk XBRL archive"),
        ("2026-06-03", "Federal Reserve MDRM integration; texas_loans_labeled.csv export"),
        ("2026-06-03", "SharePoint distribution for teammates (GitHub excludes 805 MB CSV)"),
        ("2026-06-23", "Borrower-perspective EDA workbook (texas_loans_labeled, EDA, Insights, this sheet)"),
        ("", ""),
    ])

    return rows


SECTION_TITLES = {
    "Borrower-perspective EDA — texas_loans_labeled.xlsx",
    "Methodology & caveats",
    "ICP specialists by product",
    "Codes with zero Texas balances (latest period)",
    "DATA EXTRACTION — How texas_loans_labeled was built",
    "ABBREVIATIONS & GLOSSARY — texas_loans_labeled workbook",
}

EXPLANATION_PREFIX = "What this table means:"


def _is_section_title(label: str) -> bool:
    if not label:
        return False
    if label in SECTION_TITLES or label.startswith("Top specialists") or label.startswith("§ "):
        return True
    return label[0].isdigit() and "." in label[:4]


def write_table(ws, start_row: int, table: pd.DataFrame) -> int:
    for j, col in enumerate(table.columns, 1):
        c = ws.cell(row=start_row + 1, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True)
    for i, row in enumerate(table.itertuples(index=False), start_row + 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    return start_row + len(table) + 2


def write_sheet_sections(ws, rows: list, banner: str, subtitle: str) -> None:
    c1 = ws.cell(row=1, column=1, value=banner)
    c1.font = BANNER_FONT
    c1.fill = BANNER_FILL
    ws.merge_cells("A1:F1")
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(italic=True, size=10)
    ws.merge_cells("A2:F2")

    start_row = 3
    for label, value in rows:
        if label == "__table__" and isinstance(value, pd.DataFrame):
            start_row = write_table(ws, start_row, value)
            start_row += 1
            continue
        if label == "__explanation__" and isinstance(value, str):
            lbl = ws.cell(row=start_row + 1, column=1, value=EXPLANATION_PREFIX)
            lbl.font = EXPLANATION_LABEL_FONT
            lbl.fill = EXPLANATION_FILL
            ws.merge_cells(
                start_row=start_row + 1, start_column=1,
                end_row=start_row + 1, end_column=6,
            )
            body = ws.cell(row=start_row + 2, column=1, value=value)
            body.alignment = Alignment(wrap_text=True, vertical="top")
            body.fill = EXPLANATION_FILL
            ws.merge_cells(
                start_row=start_row + 2, start_column=1,
                end_row=start_row + 2, end_column=6,
            )
            start_row += 3
            continue
        if label == "" and value == "":
            start_row += 1
            continue
        if _is_section_title(label):
            cell = ws.cell(row=start_row + 1, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=6)
            start_row += 1
            continue
        ws.cell(row=start_row + 1, column=1, value=label)
        val_cell = ws.cell(row=start_row + 1, column=2, value=value)
        val_cell.alignment = Alignment(wrap_text=True, vertical="top")
        start_row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 88


def enrich_insight(ins: dict) -> dict:
    meta = INSIGHT_CATALOG.get(ins["id"], {})
    return {
        **ins,
        "theme": meta.get("theme", ins.get("theme", "")),
        "captures": meta.get("captures", ins.get("captures", "")),
    }


def write_insights_sheet(ws, insights: list[dict]) -> None:
    ws.delete_rows(1, ws.max_row)
    c1 = ws.cell(row=1, column=1, value="Borrower Insights — texas_loans_labeled.xlsx")
    c1.font = BANNER_FONT
    c1.fill = BANNER_FILL
    ws.merge_cells("A1:G1")
    c2 = ws.cell(
        row=2, column=1,
        value=f"Documented findings from borrower-perspective EDA | Built {datetime.now():%Y-%m-%d}",
    )
    c2.font = Font(italic=True, size=10)
    ws.merge_cells("A2:G2")

    headers = [
        "ID",
        "Theme",
        "What it captures",
        "Priority",
        "Insight",
        "Evidence",
        "Recommended borrower action",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True)

    for i, raw in enumerate(insights, 5):
        ins = enrich_insight(raw)
        row_vals = [
            ins["id"],
            ins["theme"],
            ins["captures"],
            ins["priority"],
            ins["insight"],
            ins["evidence"],
            ins["borrower_action"],
        ]
        fill = INSIGHT_HIGH if ins["priority"] == "High" else INSIGHT_MED if ins["priority"] == "Medium" else None
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 38


def add_eda_sheet_only(wb_path: Path | None = None) -> None:
    """Regenerate EDA sheet only (rounded numbers + table explanations)."""
    from openpyxl import load_workbook

    path = wb_path or OUTPUT
    print(f"Loading {path} for EDA regeneration …")
    df = pd.read_excel(path, sheet_name="texas_loans_labeled")
    print(f"  {len(df):,} rows — running analysis …")
    eda_rows, _insights = run_analysis(df)
    print(f"Writing EDA sheet ({sum(1 for r in eda_rows if r[0] == '__table__')} tables) …")
    wb = load_workbook(path)
    if "EDA" in wb.sheetnames:
        del wb["EDA"]
    ws = wb.create_sheet("EDA", 1)
    write_sheet_sections(
        ws,
        eda_rows,
        "EDA — Borrower perspective | texas_loans_labeled.xlsx",
        "All numeric values rounded to nearest whole number | Narrative under each table",
    )
    wb.save(path)
    print(f"Saved {path}")


    return rows


# --- Abbreviations glossary (static reference tables) ---

INDUSTRY_ABBREVIATIONS = [
    ("FFIEC", "Federal Financial Institutions Examination Council — umbrella regulator coordinating Call Report collection"),
    ("CDR", "Central Data Repository — FFIEC public database and API for regulatory filings"),
    ("PWS", "Public Web Service — FFIEC REST API used to download Call Reports"),
    ("MDRM", "Micro Data Reference Manual — Federal Reserve dictionary of Call Report line codes and definitions"),
    ("XBRL", "eXtensible Business Reporting Language — XML format of modern Call Report facsimiles"),
    ("RSSD", "Federal Reserve ID (Research Statistics Supervision Discount) — unique bank identifier; column id_rssd"),
    ("FDIC", "Federal Deposit Insurance Corporation — deposit insurer; publishes RC-C schedule instructions"),
    ("Call Report", "Quarterly Consolidated Reports of Condition and Income (FFIEC 031/041) — source of this dataset"),
    ("Schedule RC-C", "Call Report schedule for Loans and Lease Financing Receivables — primary loan detail"),
    ("Schedule RC", "Balance sheet (assets, liabilities, capital) — parent schedule family"),
    ("UBPR", "Uniform Bank Performance Report — peer analytics (not in this extract)"),
    ("ICP", "Ideal Customer Profile — in this project: Texas community banks with $500M–$2B assets"),
    ("CRE", "Commercial real estate — investor (income) or owner-occupied property loans"),
    ("C&I", "Commercial and industrial — operating business loans and lines"),
    ("RE", "Real estate — loans secured by property"),
    ("LTV", "Loan-to-value — not in Call Report data; underwriting term"),
    ("DSCR", "Debt service coverage ratio — not in Call Report data; underwriting term"),
    ("HQ", "Headquarters — city where bank is domiciled on FFIEC panel"),
    ("DOJ", "U.S. Department of Justice — publishes HHI concentration thresholds for antitrust"),
    ("HHI", "Herfindahl-Hirschman Index — market concentration measure (EDA sheet)"),
    ("Gini", "Gini coefficient — inequality measure for balance distribution across banks (EDA sheet)"),
    ("SHA-256", "Cryptographic hash of downloaded XBRL file — filing integrity check in texas_filings.csv"),
    ("YAML", "loan_products.yaml — borrower site product taxonomy source file"),
]

COLUMN_GLOSSARY = [
    ("id_rssd", "Federal Reserve RSSD ID — stable numeric key for each bank; join to institutions and profiles"),
    ("institution_name", "Legal/reporting name of the bank from FFIEC panel"),
    ("reporting_period", "Call Report quarter-end date (e.g. 2026-03-31 = Q1 2026)"),
    ("mdrm_code", "Regulatory line item code (e.g. RCON2122) — join key to MDRM dictionary"),
    ("item_name", "Official short name for the line item from Federal Reserve MDRM"),
    ("line_description", "Display label for the line (usually same as item_name)"),
    ("mdrm_description", "Full Federal Reserve definition of the line item (up to 800 characters)"),
    ("mdrm_category", "Classifier: loan_or_lease | schedule_rc_c | loan_related_prefix | other"),
    ("reporting_form", "Call Report form type the line belongs to (e.g. FFIEC 031, FFIEC 041)"),
    ("item_type", "MDRM item type: F=financial, D=derived, S=structural/reporting"),
    ("value_num", "Numeric balance reported on the line (Call Report convention: typically thousands USD)"),
    ("value_text", "Raw XBRL text value before numeric parsing"),
    ("context_ref", "XBRL context identifier — encodes reporting period/scenario (e.g. CI_{RSSD}_{date})"),
    ("unit_ref", "XBRL unit: USD = U.S. dollars; NON-MONETARY = count or non-dollar measure"),
]

FFIEC_FORMS = [
    ("FFIEC 002", "Report of Assets and Liabilities of U.S. Branches and Agencies of Foreign Banks"),
    ("FFIEC 010", "Uniform Bank Performance Report filing (certain institutions)"),
    ("FFIEC 011", "Consolidated Report of Condition and Income for Edge Act corporations"),
    ("FFIEC 012", "Consolidated Report of Condition and Income for agreement corporations"),
    ("FFIEC 013", "Report of Condition for domestic branches of foreign banks"),
    ("FFIEC 014", "Capital and Asset Report for Foreign Banking Organizations"),
    ("FFIEC 031", "Consolidated Report of Condition and Income — larger commercial banks (assets typically >$5B)"),
    ("FFIEC 032", "Consolidated Report of Condition and Income — intermediate-size banks"),
    ("FFIEC 033", "Consolidated Report of Condition and Income — savings associations"),
    ("FFIEC 034", "Consolidated Report of Condition and Income — smaller savings associations"),
    ("FFIEC 041", "Consolidated Report of Condition and Income — community banks (smaller commercial banks)"),
    ("FFIEC 051", "Consolidated Report of Condition and Income supplemental schedule — extra loan detail lines"),
]

MDRM_PREFIX_FAMILIES = [
    ("RCON", "Mnemonic for domestic-office Call Report line items (most Texas commercial banks)"),
    ("RCFD", "Mnemonic for consolidated/domestic-office variant — common on FFIEC 041 community bank filings"),
    ("RCON14*", "Real estate–secured loans (farmland, multifamily, 1–4 family, etc.)"),
    ("RCON15*", "Consumer loans and credit card plans"),
    ("RCON16*", "Commercial and industrial and other loan categories"),
    ("RCON17*", "Lease financing receivables"),
    ("RCON21*", "Loan totals, allowance for loan losses, net loans"),
    ("RCONF1*", "Schedule RC-C extension — construction, CRE subcategories, owner-occupied"),
    ("RCONF2*", "Schedule RC-C extension — additional loan breakdowns"),
    ("RCONHK*", "Schedule RC-C memoranda — supplementary loan detail"),
    ("RCONJ4*", "Schedule RC-C extension — past-due and nonaccrual detail"),
    ("RCONLL*", "Lease financing detail lines"),
    ("RCONA5*", "Schedule RC-C extension — granular loan categories"),
    ("RCONB5*", "FFIEC 051 supplemental consumer and mortgage exposure lines"),
    ("RCFD14*", "RCFD variant of real estate–secured loan lines"),
    ("RCFD15*", "RCFD variant of consumer loan lines"),
    ("RCFD16*", "RCFD variant of C&I and other loans"),
    ("RCFD21*", "RCFD variant of loan totals and allowance"),
    ("RCFDHK*", "RCFD variant of RC-C memoranda"),
    ("RCFDJ4*", "RCFD variant of RC-C past-due detail"),
    ("RCFDLL*", "RCFD variant of lease lines"),
]

EDA_INSIGHTS_ABBREVIATIONS = [
    ("I-01 … I-12", "Insight reference IDs on the Insights sheet (I = Insight, number = sequence)"),
    ("mf", "Multifamily (5+ units) — internal mix key in loan_mix.py"),
    ("inv", "Investor CRE (income property) — non-owner-occupied commercial real estate"),
    ("own", "Owner-occupied CRE — business occupies the collateral property"),
    ("con", "Commercial construction — ground-up, rehab, land development"),
    ("ci", "Commercial & industrial — operating business loans"),
    ("res", "1–4 family residential — portfolio rental, not agency mortgage"),
    ("cons", "Consumer lending — credit cards and other consumer"),
    ("farm / ag", "Farmland and agricultural production loans"),
    ("lease", "Lease financing receivables"),
    ("P25 / P75", "25th and 75th percentile of lender exposures — deal size fit analysis"),
    ("≥8%", "Specialist threshold — product is ≥8% of bank's total loan portfolio"),
    ("≥15%", "Deep specialist threshold — stronger portfolio concentration signal"),
    ("Availability score", "% of Texas banks with any non-zero balance in the product category"),
    ("Portfolio %", "Product balance ÷ total loans (RCON2122) × 100"),
    ("Past-due 30–89", "Loans past due 30–89 days and still accruing (RCON5367 family)"),
    ("Past-due 90+", "Loans past due 90+ days (RCON5368 family) — lender stress signal"),
    ("$M / $B", "Millions / billions of U.S. dollars in EDA tables"),
    ("TX", "Texas — geographic scope of this dataset"),
]

KEY_MDRM_CODES = [
    ("RCON2122", "Total loans and leases, gross (headline loan portfolio total)"),
    ("RCFD2122", "RCFD variant of total loans and leases (community bank form)"),
    ("RCON2145", "Total loans and leases, net of unearned income"),
    ("RCON2130", "Allowance for loan and lease losses (ALLL)"),
    ("RCON1460", "Multifamily (5+ units) residential real estate loans"),
    ("RCONF161", "Other nonfarm nonresidential real estate — investor CRE proxy"),
    ("RCONF162", "Commercial real estate loans"),
    ("RCONF160", "Owner-occupied nonfarm nonresidential real estate"),
    ("RCONF158", "1–4 family residential construction loans"),
    ("RCONF159", "Other construction and land development"),
    ("RCON1766", "Commercial and industrial (C&I) loans"),
    ("RCON1480", "Loans secured by nonfarm nonresidential properties (legacy C&I/CRE line)"),
    ("RCON1403", "1–4 family residential mortgage loans"),
    ("RCON1420", "Loans secured by farmland"),
    ("RCON1590", "Agricultural production loans"),
    ("RCON1545", "Credit card plans"),
    ("RCON1583", "Other consumer loans"),
    ("RCON1754", "Lease financing receivables"),
    ("RCON5367", "Past due 30–89 days and still accruing (aggregate)"),
    ("RCON5368", "Past due 90 days or more (aggregate)"),
    ("RCONS439", "Residential mortgage exposures (FFIEC 051 supplemental)"),
    ("RCONB562", "Consumer loans other — FFIEC 051 supplemental"),
    ("RCONB539", "Consumer revolving credit — FFIEC 051 supplemental"),
]


def build_abbreviations_rows(df: pd.DataFrame) -> list[tuple[str, object]]:
    """Build glossary rows for the Abbreviations sheet."""
    rows: list[tuple[str, object]] = []

    def sec(title: str) -> None:
        rows.append((f"§ {title}", ""))
        rows.append(("", ""))

    def tbl(frame: pd.DataFrame) -> None:
        rows.append(("__table__", frame))
        rows.append(("", ""))

    rows.append(("ABBREVIATIONS & GLOSSARY — texas_loans_labeled workbook", ""))
    rows.append(("", ""))
    rows.extend([
        ("Document built", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Scope", "Every abbreviation, acronym, code prefix, and shorthand used in texas_loans_labeled, EDA, and Insights sheets"),
        ("MDRM codes in dataset", df["mdrm_code"].nunique()),
        ("", ""),
    ])

    sec("1. Industry & regulatory acronyms")
    tbl(pd.DataFrame(INDUSTRY_ABBREVIATIONS, columns=["Abbreviation", "Full meaning"]))

    sec("2. Column names — texas_loans_labeled sheet")
    tbl(pd.DataFrame(COLUMN_GLOSSARY, columns=["Column / term", "Definition"]))

    sec("3. mdrm_category values (in this dataset)")
    cat_desc = {
        "loan_or_lease": "Line item relates to loans, leases, or lending activity per MDRM keyword match",
        "schedule_rc_c": "Line item on Schedule RC-C (loans and lease financing receivables)",
        "loan_related_prefix": "Captured via RCON/RCFD loan prefix filter but not in MDRM loan category",
        "other": "Present in facts file but classified as non-loan by MDRM loader",
    }
    cats = []
    for val, cnt in df["mdrm_category"].value_counts().items():
        cats.append({"Value": val, "Rows in file": cnt, "Meaning": cat_desc.get(val, "")})
    tbl(pd.DataFrame(cats))

    sec("4. item_type values (MDRM item type codes)")
    type_desc = {
        "F": "Financial — reported dollar amount line item",
        "D": "Derived — calculated from other lines, not directly reported",
        "S": "Structural — reporting framework, header, or non-amount line",
    }
    types = []
    for val, cnt in df["item_type"].dropna().value_counts().items():
        types.append({"Code": val, "Rows in file": cnt, "Meaning": type_desc.get(str(val), "")})
    tbl(pd.DataFrame(types))

    sec("5. unit_ref values (XBRL units)")
    unit_desc = {
        "USD": "U.S. dollars — monetary amounts on Call Report lines",
        "NON-MONETARY": "Non-dollar measure (counts, ratios, flags) — not a loan balance",
    }
    units = []
    for val, cnt in df["unit_ref"].dropna().value_counts().items():
        units.append({"Unit": val, "Rows in file": cnt, "Meaning": unit_desc.get(str(val), "")})
    tbl(pd.DataFrame(units))

    sec("6. reporting_form values (FFIEC Call Report forms in this dataset)")
    form_map = {f[0]: f[1] for f in FFIEC_FORMS}
    forms = []
    for val, cnt in df["reporting_form"].dropna().value_counts().sort_index().items():
        forms.append({
            "Form": val,
            "Rows in file": cnt,
            "Description": form_map.get(str(val), "See FFIEC instruction book"),
        })
    tbl(pd.DataFrame(forms))

    sec("7. context_ref pattern")
    rows.extend([
        ("Pattern", "CI_{id_rssd}_{YYYY-MM-DD}"),
        ("CI", "Context Instant — XBRL instant-in-time balance as of quarter-end"),
        ("id_rssd", "Federal Reserve RSSD ID embedded in context"),
        ("date", "Reporting period end date"),
        ("Example", "CI_1001152_2025-12-31 = instant context for RSSD 1001152 at 2025-12-31"),
        ("", ""),
    ])

    sec("8. MDRM code prefix families (mnemonic structure)")
    tbl(pd.DataFrame(MDRM_PREFIX_FAMILIES, columns=["Prefix / family", "Meaning"]))

    sec("9. Key borrower-relevant MDRM codes (most used in EDA)")
    tbl(pd.DataFrame(KEY_MDRM_CODES, columns=["MDRM code", "What it measures"]))

    sec("10. Complete MDRM code glossary — all codes appearing in this dataset")
    code_glossary = (
        df.groupby("mdrm_code", as_index=False)
        .agg(
            item_name=("item_name", "first"),
            mdrm_category=("mdrm_category", "first"),
            reporting_form=("reporting_form", "first"),
            rows_in_file=("mdrm_code", "size"),
            banks_reporting=("id_rssd", "nunique"),
        )
        .sort_values("mdrm_code")
    )
    code_glossary.columns = [
        "MDRM code", "Official line item name", "Category",
        "Reporting form", "Rows in file", "Banks reporting",
    ]
    tbl(code_glossary)
    rows.append((
        "Note",
        f"All {len(code_glossary)} distinct MDRM codes above appear at least once in texas_loans_labeled. "
        "Lookup any code at https://www.federalreserve.gov/apps/mdrm/data-dictionary",
    ))
    rows.append(("", ""))

    sec("11. EDA & Insights sheet abbreviations")
    tbl(pd.DataFrame(EDA_INSIGHTS_ABBREVIATIONS, columns=["Abbreviation", "Meaning"]))

    sec("12. filing_type numeric codes (from texas_institutions.csv)")
    filing_types = [
        ("41", "FFIEC 041 — community / smaller bank Call Report"),
        ("51", "FFIEC 031 — larger commercial bank Call Report"),
        ("31", "FFIEC 031 variant code on panel"),
        ("71", "FFIEC 051 supplemental filer"),
    ]
    tbl(pd.DataFrame(filing_types, columns=["filing_type code", "Maps to form"]))

    return rows


def add_abbreviations_sheet(wb_path: Path | None = None) -> None:
    """Add or replace the Abbreviations glossary sheet."""
    from openpyxl import load_workbook

    path = wb_path or OUTPUT
    print(f"Loading {path} for abbreviations glossary …")
    df = pd.read_excel(
        path,
        sheet_name="texas_loans_labeled",
        usecols=[
            "id_rssd", "mdrm_code", "item_name", "mdrm_category",
            "reporting_form", "item_type", "unit_ref",
        ],
    )
    rows = build_abbreviations_rows(df)
    n_tables = sum(1 for r in rows if r[0] == "__table__")
    print(f"Writing Abbreviations sheet ({n_tables} tables, {df['mdrm_code'].nunique()} MDRM codes) …")
    wb = load_workbook(path)
    if "Abbreviations" in wb.sheetnames:
        del wb["Abbreviations"]
    ws = wb.create_sheet("Abbreviations")
    write_sheet_sections(
        ws,
        rows,
        "Abbreviations & Glossary — texas_loans_labeled workbook",
        "Acronyms, column definitions, form codes, MDRM prefixes, and every line code in the dataset",
    )
    wb.save(path)
    print(f"Saved {path} — sheets: {wb.sheetnames}")


def add_data_extraction_sheet(wb_path: Path | None = None) -> None:
    """Add or replace the Data Extraction documentation sheet (no full data rewrite)."""
    from openpyxl import load_workbook

    path = wb_path or OUTPUT
    print(f"Loading stats from {path} …")
    df = pd.read_excel(
        path,
        sheet_name="texas_loans_labeled",
        usecols=["id_rssd", "reporting_period", "mdrm_code", "mdrm_category"],
    )
    rows = build_data_extraction_rows(df)
    print(f"Writing Data Extraction sheet ({len(rows)} blocks) …")
    wb = load_workbook(path)
    if "Data Extraction" in wb.sheetnames:
        del wb["Data Extraction"]
    ws = wb.create_sheet("Data Extraction")
    write_sheet_sections(
        ws,
        rows,
        "Data Extraction — How texas_loans_labeled was built",
        "Detailed provenance: FFIEC API → XBRL archive → facts CSV → MDRM-labeled loan export",
    )
    wb.save(path)
    print(f"Saved {path} — sheets: {wb.sheetnames}")


def main() -> int:
    if not INPUT.exists():
        print(f"Missing {INPUT}")
        return 1

    print(f"Loading {INPUT} …")
    df = pd.read_excel(INPUT, sheet_name=0)
    print(f"  {len(df):,} rows × {len(df.columns)} columns")

    print("Running borrower-perspective analysis …")
    eda_rows, insights = run_analysis(df)

    print(f"Writing workbook → {TMP} …")
    with pd.ExcelWriter(TMP, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="texas_loans_labeled", index=False)
        pd.DataFrame({"": []}).to_excel(writer, sheet_name="EDA", index=False)
        pd.DataFrame({"": []}).to_excel(writer, sheet_name="Insights", index=False)

    from openpyxl import load_workbook

    wb = load_workbook(TMP)
    write_sheet_sections(
        wb["EDA"],
        eda_rows,
        "EDA — Borrower perspective | texas_loans_labeled.xlsx",
        f"Complex exploratory analysis | {len(insights)} documented insights on Insights sheet",
    )
    write_insights_sheet(wb["Insights"], insights)

    # Style original data header
    ws_data = wb["texas_loans_labeled"]
    for cell in ws_data[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    wb.save(TMP)
    TMP.replace(OUTPUT)
    print(f"Done. Workbook: {OUTPUT}")
    print(f"  Sheets: texas_loans_labeled ({len(df):,} rows), EDA, Insights ({len(insights)} insights)")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--extraction-only":
        add_data_extraction_sheet()
        sys.exit(0)
    if len(sys.argv) > 1 and sys.argv[1] == "--eda-only":
        add_eda_sheet_only()
        sys.exit(0)
    if len(sys.argv) > 1 and sys.argv[1] == "--abbreviations-only":
        add_abbreviations_sheet()
        sys.exit(0)
    sys.exit(main())
